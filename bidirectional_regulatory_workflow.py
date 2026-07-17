#!/usr/bin/env python3
"""Build a bidirectional regulation <-> management-action review workpaper.

Forward chain:
    source file -> clause -> atomic obligation -> standard management action

Reverse chain:
    standard management action -> execution-plan task -> cited clauses

The merge operation never discards the obligation ID or clause ID.  Coverage
gaps, unresolved classifications and standard actions not represented in the
execution plan are written to a dedicated worksheet.
"""

from __future__ import annotations

import json
import re
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import regulatory_basis_review as base


ROOT = base.ROOT
OUTPUT = ROOT / "制度依据双向复合复核工作底稿-授信系统群咨询项目.xlsx"
DATA_DIR = base.DATA_DIR


@dataclass
class ActionRule:
    action_id: str
    category: str
    title: str
    action_type: str
    default_applicability: str
    task_attribute: str
    keywords: list[str]
    description: str


@dataclass
class Obligation:
    obligation_id: str
    source_id: str
    clause_id: str
    clause_title: str
    atomic_text: str
    subject: str
    object_: str
    action_summary: str
    timing: str
    evidence: str
    applicability: str
    relation: str
    action_ids: list[str]
    classification_status: str
    note: str = ""


ACTION_RULES = [
    ActionRule("CTX_01", "00制度基线与适用性", "维护制度基线、效力状态和适用性矩阵", "治理基础", "直接适用", "持续/周期任务", ["适用范围", "制定本办法", "分类管理机制", "制度和流程", "监管政策", "效力", "法律法规"], "识别制度适用范围、效力和项目触发条件。"),
    ActionRule("CTX_02", "00制度基线与适用性", "记录监管机关职责及无需项目执行的条款", "非项目执行", "机构外部/不形成项目任务", "基线记录", ["监督管理", "监管机构", "银保监会及其派出机构", "金融监管总局及", "负责解释", "行业自律"], "监管机关、解释或行业推动职责，不直接转化为项目组动作。"),
    ActionRule("GOV_01", "01治理与职责", "建立外包治理架构并明确职责边界", "监管管理动作", "机构层面适用（项目配合）", "持续任务", ["治理架构", "职责", "董事会", "高级管理层", "管理委员会", "牵头部门", "责任制"], "明确治理主体、风险管理、执行团队和协同部门职责。"),
    ActionRule("GOV_02", "01治理与职责", "落实主体责任并保留核心管理能力", "监管管理动作", "直接适用", "持续任务", ["不得外包", "主体责任", "核心竞争力", "核心能力", "自主可控", "最低限度的服务能力"], "管理责任、安全主体责任和核心能力不得随外包转移。"),
    ActionRule("STR_01", "02战略分类与项目判断", "制定和执行信息科技外包战略", "机构管理动作", "机构层面适用（项目遵循）", "周期任务", ["外包战略", "外包原则", "资源能力建设", "风险偏好"], "机构制定外包战略，项目验证一致性。"),
    ActionRule("CLS_01", "02战略分类与项目判断", "识别外包合作模式和服务类型", "监管管理动作", "直接适用", "一次性+变更复核", ["咨询规划类", "开发测试类", "运行维护类", "安全服务类", "业务支持类", "合作模式", "服务类型", "分类"], "确认项目外包类型、合作模式和配套制度。"),
    ActionRule("CLS_02", "02战略分类与项目判断", "判断重要外包并实施差异化管理", "监管管理动作", "直接判断、结论条件适用", "一次性+变更复核", ["重要外包", "一般外包", "分级管理", "差异化管控"], "每个项目完成重要外包判断并承接强化管理。"),
    ActionRule("CLS_03", "02战略分类与项目判断", "判断重大信息科技项目并承接治理要求", "监管管理动作", "直接判断、结论条件适用", "一次性+周期/后续", ["重大信息科技项目", "重大项目", "投产后", "后评价"], "完成重大项目判断，触发信科委报告和投产后评价。"),
    ActionRule("RSK_01", "03风险评估与决策", "开展项目外包风险评估和实施决策", "监管管理动作", "直接适用", "一次性+变更触发", ["拟开展", "风险评估", "识别并评估", "审慎决策", "一致性", "风险及防控措施"], "外包前和重大变化后评估风险并作出决策。"),
    ActionRule("RSK_02", "03风险评估与决策", "开展年度全面外包风险评估", "机构管理动作", "机构层面适用（项目提供材料）", "周期任务", ["每年应当至少开展一次全面", "年度风险评估", "全面的信息科技外包风险"], "机构每年至少一次全面评估并向治理层报告。"),
    ActionRule("RSK_03", "03风险评估与决策", "识别并降低外包集中度风险", "监管管理动作", "直接适用", "选型+周期任务", ["集中度风险", "个别外包服务提供商", "替代服务提供商", "分散外包"], "识别单一服务商及行业集中度风险并落实缓释。"),
    ActionRule("VEN_01", "04供应商准入与尽调", "制定供应商准入标准并实施筛选", "监管管理动作", "直接适用", "一次性+变更复核", ["准入标准", "备选服务提供商", "筛选", "准入审核"], "基于风险和外包战略设置准入标准。"),
    ActionRule("VEN_02", "04供应商准入与尽调", "开展供应商尽职调查", "监管管理动作", "重要外包直接/其他按行内制度", "一次性+有效期复核", ["尽职调查", "经营状况", "技术和行业经验", "内部控制和管理能力", "调查材料"], "合同前调查能力、经营、内控、安全、关联性等。"),
    ActionRule("VEN_03", "04供应商准入与尽调", "实施非驻场、跨境和关联外包专项审查", "条件管理动作", "条件适用", "一次性+变更复核", ["非驻场", "跨境外包", "关联外包", "所在国家或地区", "系统和数据是否有明确"], "对应特殊外包形态开展强化调查或利益冲突审查。"),
    ActionRule("VEN_04", "04供应商准入与尽调", "维护外包商分类分级和强化管理", "行内管理动作", "直接判断、结论条件适用", "持续/周期任务", ["一级外包商", "二级外包商", "外包商分类分级", "管理要求按照级别"], "形成分类分级台账并实施差异化管理。"),
    ActionRule("CON_01", "05合同与分包管理", "明确合同服务范围、交付物、责任和质量约定", "监管+项目管理动作", "直接适用", "合同前+变更触发", ["服务范围", "服务内容", "服务要求", "工作时限", "责任分配", "交付物要求", "服务质量考核"], "合同明确服务边界、交付、人员、质量和责任。"),
    ActionRule("CON_02", "05合同与分包管理", "保留风险评估、检查、审计和监管延伸检查权", "监管管理动作", "直接适用", "合同前+变更触发", ["风险评估、监测、检查和审计", "监督检查", "审计的权利", "延伸检查"], "合同保留银行和监管穿透检查权。"),
    ActionRule("CON_03", "05合同与分包管理", "明确合规、安全保密、数据及报告条款", "监管管理动作", "直接适用", "合同前+变更触发", ["安全保密", "消费者权益", "报告条款", "数据以任何形式转移", "合同允许范围外", "保护义务"], "合同明确合规、安全、数据使用禁止和报告路线。"),
    ActionRule("CON_04", "05合同与分包管理", "限制转包分包并穿透管理分包商", "监管管理动作", "直接适用", "合同前+分包变更触发", ["不得将外包服务转包", "分包", "主要业务", "主服务提供商对服务水平负总责"], "禁止转包并对必要分包设置限制和变更审批。"),
    ActionRule("CON_05", "05合同与分包管理", "履行信息科技风险、法律和信科委合同审核", "监管管理动作", "直接适用", "合同签订/重大变更前", ["所有信息科技外包合同", "法律部门", "信息科技管理委员会审核", "合同审核", "会签"], "按2009年第62条和行内流程完成三方审核。"),
    ActionRule("CON_06", "05合同与分包管理", "管理合同变更、终止和过渡安排", "监管管理动作", "直接适用/条件触发", "合同前约定+变化触发", ["合同变更或终止", "过渡安排", "终止交接", "合同意外终止", "退出策略"], "事前约定触发条件，变化后评估并履行审核。"),
    ActionRule("SLA_01", "06服务目录与质量监控", "建立并维护信息科技外包服务目录", "监管管理动作", "直接适用", "建立后持续更新", ["外包服务目录", "服务目录"], "维护服务范围、阶段和服务提供商目录。"),
    ActionRule("SLA_02", "06服务目录与质量监控", "建立、定期审阅和修订SLA", "监管管理动作", "直接适用", "合同前建立+定期审阅+触发修订", ["服务水平协议", "审阅和修订服务水平协议", "SLA"], "咨询项目采用交付质量、人员、整改、监理和售后指标。"),
    ActionRule("SLA_03", "06服务目录与质量监控", "建立服务效能和质量指标并监控", "监管管理动作", "直接适用", "服务期间按频率监控", ["服务效能和质量监控指标", "服务质量指标", "客户满意度", "及时完成率", "考核合格率"], "建立量化口径、目标值、数据源、频率和未达标处理。"),
    ActionRule("SLA_04", "06服务目录与质量监控", "持续监控服务并整改异常", "监管管理动作", "直接适用", "服务期间持续", ["持续监控", "异常情况", "纠正措施", "限期整改", "约谈", "服务质量的急剧下降"], "监控外包过程、服务商经营内控安全和异常整改。"),
    ActionRule("PER_01", "07人员与驻场管理", "实施人员入场、资质背景和变更管理", "行内管理动作", "存在驻场时直接适用", "入场前+变更触发", ["入场申请", "资质", "背景调查", "劳动关系", "人员变更", "简历"], "核验身份资质、劳动关系、背景及人员替换。"),
    ActionRule("PER_02", "07人员与驻场管理", "开展安全培训并签署保密承诺", "监管+行内动作", "直接适用", "入场前+新增/变化触发", ["安全教育或培训", "安全保密协议", "保密承诺书", "宣贯培训", "考试"], "人员入场及变化时完成培训、考试和保密承诺。"),
    ActionRule("PER_03", "07人员与驻场管理", "实施必需知道和最小授权的访问控制", "监管管理动作", "发生访问时直接适用", "授权前+周期复核+离场注销", ["必需知道", "最小授权", "访问授权", "访问权限", "远程维护", "用户认证和访问控制"], "建立权限申请、审批、配置、复核和注销闭环。"),
    ActionRule("PER_04", "07人员与驻场管理", "管理驻场考勤、行为、设备和办公安全", "行内管理动作", "存在驻场时直接适用", "驻场期间持续", ["考勤", "行为规范", "办公设备", "消防安全", "工位", "门禁", "现场服务"], "落实驻场日常纪律、设备、工位和异常排查。"),
    ActionRule("DAT_01", "08数据安全与个人信息", "识别数据处理模式并开展数据分类分级", "监管管理动作", "发生数据活动时直接适用", "处理前+变化复核", ["数据分类分级", "数据目录", "数据级别", "委托处理", "共同处理", "数据处理活动"], "识别查看、提取、委托、共同处理、共享和传输模式。"),
    ActionRule("DAT_02", "08数据安全与个人信息", "开展数据安全评估和处理前置判断", "监管管理动作", "条件适用", "敏感级处理或高影响活动前", ["事先开展数据安全评估", "必要性、合规性", "数据安全评估", "处理敏感级及以上"], "在敏感级处理或高影响活动前完成评估。"),
    ActionRule("DAT_03", "08数据安全与个人信息", "签署数据委托或共同处理协议并监督受托方", "监管管理动作", "条件适用", "首次处理前+处理期间", ["委托处理数据", "共同处理", "合同协议方式约定", "受托方", "不得转委托", "不得加工、训练、挪用"], "明确目的、期限、方式、范围、安全责任及删除返还。"),
    ActionRule("DAT_04", "08数据安全与个人信息", "实施数据访问审批、日志和审计", "监管管理动作", "条件适用", "访问前+持续记录+周期审计", ["数据访问", "日志记录", "操作日志", "审计周期", "生产环境提取数据"], "对敏感数据访问建立闭环、日志及周期审计。"),
    ActionRule("DAT_05", "08数据安全与个人信息", "落实数据传输、存储、脱敏和测试隔离控制", "监管管理动作", "条件适用", "处理设计/实施/投产前", ["安全的传输方式", "安全存储", "脱敏", "测试环境", "生产系统隔离", "信息系统开发"], "在系统和方案中落实传输、存储、脱敏、测试隔离。"),
    ActionRule("DAT_06", "08数据安全与个人信息", "返还删除销毁数据并验证不可恢复", "监管管理动作", "条件适用", "期限届满/终止/离场时", ["删除或者销毁", "不可恢复", "返还或者删除数据", "介质报废", "销毁"], "服务结束或使用期满时完成返还、删除和不可恢复验证。"),
    ActionRule("DAT_07", "08数据安全与个人信息", "开展个人信息影响评估并落实保护义务", "监管管理动作", "条件适用", "处理前+处理期间", ["个人信息保护影响评估", "个人敏感", "个人信息", "数据主体", "合法权益"], "识别个人信息场景，开展影响评估并监督委托处理。"),
    ActionRule("DAT_08", "08数据安全与个人信息", "管理数据收集、外部采购和数据服务", "监管管理动作", "条件适用", "收集/采购/服务前后持续", ["数据收集", "收集数据", "外部数据采购", "合作引入", "数据服务", "数据来源", "数据加工"], "管理数据收集必要性、外部数据引入、来源合法性和数据服务。"),
    ActionRule("DAT_09", "08数据安全与个人信息", "管理数据共享、对外提供、转移、公开和出境", "监管管理动作", "条件适用", "相关活动实施前+过程留痕", ["数据共享", "向外部提供", "数据转移", "公开", "向境外提供", "跨主体流动", "外联平台", "数据交互"], "对共享、提供、转移、公开及出境分别履行同意、评估、审批和安全控制。"),
    ActionRule("DAT_10", "08数据安全与个人信息", "建设并落实数据安全技术保护体系", "机构技术管理动作", "机构层面适用（项目方案承接）", "规划建设+持续运行", ["数据安全技术", "技术保护体系", "数据基础设施", "备份", "安全域", "网络安全等级保护", "大数据平台", "生命周期"], "建设数据安全架构、保护基线、备份恢复、平台和系统生命周期控制。"),
    ActionRule("DAT_11", "08数据安全与个人信息", "持续监测数据安全风险并开展检查评估", "机构管理动作", "机构层面/项目配合", "持续监测+年度评估+周期审计", ["数据安全威胁", "风险监测", "主动评估", "监督检查", "每年开展一次数据安全", "数据安全全面审计"], "监测异常访问、流动、外包处理和泄露风险，开展评估审计。"),
    ActionRule("AI_01", "09人工智能与模型", "识别AI实际使用和方案场景并建立清单", "监管管理动作", "直接识别、场景条件适用", "首次识别+变化复核", ["人工智能应用场景", "应用清单", "场景", "生成式人工智能", "人工智能开发应用"], "区分日常工具、行内模型和未来方案设计。"),
    ActionRule("AI_02", "09人工智能与模型", "实施AI风险分类分级和高风险准入", "监管管理动作", "条件适用", "应用前+变化复核", ["风险分类分级", "高风险应用", "风险管理委员会批准", "准入管理"], "按场景重要性、影响和模型依赖分级，高风险应用批准后实施。"),
    ActionRule("AI_03", "09人工智能与模型", "开展模型算法评审、监测、干预和退出", "监管管理动作", "条件适用", "投入使用前+运行期间", ["模型算法", "算法评估", "运行监测", "人工监督", "紧急停用", "模型退出", "测评"], "评审模型合理性、性能和安全，运行中监测并保留人工干预。"),
    ActionRule("AI_04", "09人工智能与模型", "管理外部模型、开源技术和AI供应链", "监管管理动作", "条件适用", "引入前+版本变化+定期排查", ["外部模型", "开源技术", "开源组件", "供应链", "外包合作机构", "名单制", "算力"], "登记版本许可证，评估外部模型、开源组件和集中度。"),
    ActionRule("AI_05", "09人工智能与模型", "落实AI透明、可解释、公平伦理和追溯", "监管管理动作", "条件适用", "设计/评审/运行期间", ["透明度", "可解释", "公平", "伦理", "推理路径", "人工复核", "生成内容", "责任可追溯"], "方案和系统落实透明可解释、公平伦理、人工复核和记录。"),
    ActionRule("DEL_01", "10交付与验收", "建立交付物基线、阶段评审和验收闭环", "项目管理动作", "直接适用", "建立后持续+阶段/最终验收", ["交付物", "阶段评审", "验收", "付款", "需求变更率", "及时完成率"], "由采购合同直接产生，监管质量条款提供支持。"),
    ActionRule("DEL_02", "10交付与验收", "开展交付物安全、数据和模型检查", "监管+项目动作", "按交付物内容条件适用", "交付评审/验收前", ["开发交付物", "安全扫描", "源代码", "敏感信息", "安全检查", "数据与模型"], "检查代码、脚本、文档、数据和模型风险。"),
    ActionRule("DEL_03", "10交付与验收", "完成知识转移、培训和文档交接", "监管+项目动作", "直接适用", "关键阶段+收尾", ["知识产权", "知识转移", "培训计划", "文档交接", "内部配置相应的人力资源"], "保留内部承接能力，完成文档、培训和知识移交。"),
    ActionRule("BCP_01", "11连续性、应急与事件", "建立重要外包业务连续性和应急保障", "监管管理动作", "重要且影响连续性时条件适用", "外包前+服务期间", ["业务连续性", "灾难恢复", "应急和灾备资源", "服务中断", "供应链安全保障方案"], "重要外包事先制定控制、缓释、退出和最低接管能力。"),
    ActionRule("BCP_02", "11连续性、应急与事件", "组织服务商参与应急预案和演练", "监管管理动作", "重要且影响连续性时条件适用", "预案编制+至少年度", ["参与应急计划", "应急演练", "每年在综合性演练", "定期开展应急"], "按机构统筹将适用服务商纳入预案和年度演练。"),
    ActionRule("INC_01", "11连续性、应急与事件", "建立事件联络、分级和监管报告机制", "监管管理动作", "直接建立、事件触发", "事前建立+事件后按时限", ["重大风险事件", "报告机制", "24小时内", "2小时内", "正式书面报告", "报告路线"], "建立常规和突发报告路线，事件后按监管时限报告。"),
    ActionRule("INC_02", "11连续性、应急与事件", "按场景开展外包突发事件处置和恢复", "行内管理动作", "条件适用", "事件发生后", ["应急响应阶段流程", "重大资源损失", "重大财务损失", "劳动纠纷", "供应链安全攻击", "业务/资源恢复", "处置执行结果"], "按应急预案完成启动、处置、恢复、总结和改进。"),
    ActionRule("AUD_01", "12评估审计与治理报告", "开展外包审计、检查和整改", "机构管理动作", "机构层面适用（项目配合）", "定期/三年覆盖/事件后", ["审计工作", "专项审计", "实地检查", "三年覆盖", "内部审计"], "定期审计重要外包，重大事件后专项审计。"),
    ActionRule("REP_01", "12评估审计与治理报告", "履行重要外包事前监管报告", "监管管理动作", "条件适用", "合同签订前二十个工作日", ["合同签订前二十个工作日", "监管报告", "向银保监会或其派出机构报告"], "符合列明情形时履行事前报告。"),
    ActionRule("REP_02", "12评估审计与治理报告", "向治理主体报告项目和外包风险", "机构管理动作", "机构层面/重大项目条件适用", "定期/年度/重大事项时", ["提交评估报告", "向董事会", "向高级管理层", "进度报告", "信息科技管理委员会"], "项目提供材料，治理主体完成审议审核。"),
    ActionRule("REP_03", "12评估审计与治理报告", "履行数据安全目录、活动和年度监管报告", "监管管理动作", "条件/周期适用", "活动前或按年度时限", ["重要数据目录", "批量敏感级", "数据安全风险评估报告", "处理、合同签署前二十个工作日", "报送上一年度"], "按数据安全办法报送目录、批量敏感级活动和年度报告。"),
    ActionRule("EXT_01", "13评价退出与档案", "开展履约、到期和服务商后评价", "监管+行内动作", "直接适用", "合同期间/到期前/结束时", ["是否继续外包", "服务提供商进行评价", "履约评价", "项目后评价", "合同到期"], "合同期间评价履约，到期前决策续包，结束时后评价。"),
    ActionRule("EXT_02", "13评价退出与档案", "制定退出交接计划并完成服务移交", "监管管理动作", "条件适用/结束时适用", "终止、更换或结束前后", ["退出和交接计划", "终止交接安排", "外包商退出", "业务影响分析", "服务移交"], "分析退出影响，完成交接、权限回收和替代安排。"),
    ActionRule("ARC_01", "13评价退出与档案", "归档监控评价和合规证据并执行保存期限", "监管+行内动作", "直接适用", "形成后持续归档+期限保存", ["保存到服务结束后三年", "保存时间", "归档", "记录", "真实性和完整性"], "按证据类型、期限和依据归档。"),
    ActionRule("ACC_01", "14问责整改", "落实违规责任追究和整改闭环", "监管+行内动作", "条件适用", "违规/异常发生后", ["责任追究", "问责", "处罚", "取消其服务资格", "违约", "整改闭环"], "对违规、逾期整改和重大事件落实问责处罚。"),
    ActionRule("OPS_01", "15运行服务管理", "开展信息科技运行服务水平管理和考核", "机构运行管理动作", "本咨询项目通常不直接适用", "周期任务", ["信息科技运行服务水平", "运行服务水平", "运行维护"], "2009年第45条对应运行服务考核，不作为咨询外包SLA直接依据。"),
    ActionRule("ACT_PENDING", "99待人工归类", "待人工归类的制度义务或项目任务", "待分类", "待判断", "待判断", [], "自动规则不能可靠分类时保留原条款和义务ID，禁止静默丢失。"),
]


ACTION_INDEX = {action.action_id: action for action in ACTION_RULES}


TASK_ACTION_OVERRIDES = {
    2: ["CTX_01"],
    3: ["CTX_01", "CLS_01"],
    4: ["DAT_08"],
    5: ["CLS_01"],
    11: ["GOV_02"],
    12: ["RSK_01"],
    13: ["RSK_01", "SLA_04"],
    17: ["GOV_01", "STR_01"],
    18: ["GOV_01"],
    19: ["CLS_01"],
    42: ["SLA_01", "ARC_01"],
    76: ["RSK_01", "ARC_01"],
    80: ["INC_01"],
    82: ["PER_01"], 83: ["PER_01"], 85: ["PER_01"],
    87: ["PER_02", "PER_04"],
    95: ["PER_04"],
    126: ["PER_01", "EXT_02"],
    131: ["CTX_01"],
    132: ["CLS_02", "VEN_04"],
    144: ["CON_05", "ARC_01"],
    152: ["PER_04"],
    167: ["RSK_01", "SLA_04"],
    181: ["EXT_01"],
    189: ["GOV_01", "PER_03"],
    197: ["PER_01", "CON_01"], 198: ["PER_01", "CON_01"],
    199: ["PER_01"], 200: ["PER_01"],
}


CLAUSE_ACTION_OVERRIDES = {
    # 2009 clauses with confirmed text.
    ("2009", "第32条"): ["CLS_03", "REP_02"],
    ("2009", "第45条"): ["OPS_01"],
    ("2009", "第62条"): ["CON_05", "SLA_02"],
    # Complete 141 main-line mapping.
    **{("141", f"第{i}条"): actions for i, actions in {
        1: ["CTX_01"], 2: ["CTX_01"], 3: ["CTX_01"], 4: ["CTX_01"], 5: ["GOV_02"],
        6: ["GOV_01"], 7: ["GOV_01", "REP_02"], 8: ["GOV_01", "REP_02"], 9: ["GOV_01"],
        10: ["STR_01"], 11: ["GOV_02"], 12: ["CLS_01"], 13: ["CLS_02"], 14: ["CON_06", "EXT_02"],
        15: ["RSK_01"], 16: ["VEN_01"], 17: ["VEN_02"], 18: ["VEN_03", "VEN_02"],
        19: ["VEN_03", "DAT_01"], 20: ["VEN_03"], 21: ["CON_01", "CON_02", "CON_03", "CON_06"],
        22: ["CON_04"], 23: ["SLA_04"], 24: ["SLA_01", "SLA_02", "ARC_01"], 25: ["SLA_03"],
        26: ["SLA_04"], 27: ["SLA_04", "ACC_01"], 28: ["VEN_03", "ACC_01"],
        29: ["EXT_01", "EXT_02"], 30: ["RSK_01"], 31: ["BCP_01", "BCP_02", "EXT_02"],
        32: ["PER_02", "PER_03", "DEL_02", "AI_03"], 33: ["RSK_03"], 34: ["AUD_01"],
        35: ["RSK_02", "REP_02"], 36: ["AUD_01"], 37: ["REP_01"], 38: ["INC_01"],
        39: ["CTX_02"], 40: ["CTX_02"], 41: ["CTX_02"], 42: ["CTX_02"], 43: ["CTX_02"],
        44: ["CTX_02"], 45: ["CTX_02"], 46: ["CTX_01"],
    }.items()},
    # Complete data-security mapping.  It deliberately includes institution-
    # level and regulator-duty articles so the forward completeness check does
    # not silently discard them.
    **{("DATA", f"第{i}条"): actions for i, actions in {
        1: ["CTX_01"], 2: ["CTX_01"], 3: ["CTX_01"], 4: ["CTX_02"],
        5: ["GOV_01", "DAT_10", "DAT_11"], 6: ["GOV_02"], 7: ["CTX_01", "DAT_10"],
        8: ["RSK_01", "AI_05"], 9: ["GOV_01"], 10: ["GOV_01"], 11: ["GOV_01"],
        12: ["GOV_01"], 13: ["AUD_01", "ACC_01"], 14: ["DAT_10", "DAT_11"],
        15: ["PER_02"], 16: ["DAT_01", "DAT_10"], 17: ["DAT_01", "DAT_10"],
        18: ["DAT_01", "DAT_10"], 19: ["DAT_01", "DAT_10"], 20: ["DAT_01", "DAT_10"],
        21: ["DAT_01", "DAT_10"], 22: ["DAT_02"], 23: ["DAT_08"],
        24: ["DAT_08", "DAT_07"], 25: ["DAT_08", "DAT_07"], 26: ["DAT_08", "VEN_01"],
        27: ["DAT_01", "DAT_10"], 28: ["PER_03", "DAT_04"], 29: ["DAT_09", "DAT_02"],
        30: ["DAT_03"], 31: ["DAT_03"], 32: ["DAT_03"],
        33: ["DAT_09"], 34: ["DAT_09"], 35: ["DAT_09"], 36: ["DAT_09"],
        37: ["DAT_10", "BCP_01"], 38: ["DAT_06"], 39: ["DAT_10"], 40: ["DAT_10"],
        41: ["DAT_10"], 42: ["DAT_10"], 43: ["DAT_10", "DAT_04", "ARC_01"],
        44: ["DAT_10"], 45: ["DAT_10"], 46: ["DAT_10", "DAT_06"],
        47: ["DAT_10"], 48: ["DAT_10", "DEL_02"], 49: ["DAT_10"],
        50: ["AI_03", "AI_05"], 51: ["AI_03", "AI_05"], 52: ["AI_03", "AI_05"],
        53: ["DAT_09", "DAT_05"], 54: ["DAT_07"], 55: ["DAT_07"],
        56: ["DAT_07"], 57: ["DAT_07"], 58: ["DAT_07", "DAT_02", "ARC_01"],
        59: ["DAT_07", "DAT_09"], 60: ["DAT_07", "DAT_09"],
        61: ["DAT_07", "DAT_03", "INC_01"], 62: ["DAT_07", "DAT_03", "INC_01"],
        63: ["DAT_07", "DAT_03", "INC_01"], 64: ["RSK_01", "INC_01"],
        65: ["DAT_11", "SLA_04"], 66: ["DAT_11", "RSK_02", "AUD_01"],
        67: ["INC_01"], 68: ["INC_01", "BCP_02"], 69: ["INC_01", "BCP_02"],
        70: ["CTX_02"], 71: ["DAT_01", "REP_03"], 72: ["CTX_02"],
        73: ["REP_03", "DAT_03"], 74: ["REP_03", "RSK_02"], 75: ["CTX_02"],
        76: ["CTX_02", "ACC_01"], 77: ["CTX_02", "ACC_01"], 78: ["CTX_02"],
        79: ["CTX_01"], 80: ["CTX_01"], 81: ["CTX_01"],
    }.items()},
    # Complete in-bank 3.0 mapping.
    **{("GZ3", f"第{i}条"): actions for i, actions in {
        1: ["CTX_01"], 2: ["CTX_01"], 3: ["CTX_01"], 4: ["CTX_01", "DAT_08"],
        5: ["GOV_02", "RSK_01"], 6: ["GOV_01", "STR_01", "CTX_01"],
        7: ["GOV_01", "STR_01", "CTX_01"], 8: ["GOV_01", "STR_01", "CTX_01"],
        9: ["GOV_01", "STR_01", "CTX_01"], 10: ["GOV_01", "REP_02"],
        11: ["GOV_01", "REP_02"], 12: ["GOV_01", "REP_02"],
        13: ["GOV_01", "REP_02"], 14: ["GOV_01", "REP_02", "AUD_01"],
        15: ["GOV_01", "REP_02"], 16: ["CLS_01"], 17: ["CLS_01"],
        18: ["CLS_02", "VEN_04"], 19: ["CLS_02", "VEN_04"],
        20: ["RSK_01"], 21: ["RSK_01"], 22: ["VEN_01", "RSK_03"],
        23: ["VEN_02"], 24: ["VEN_03", "VEN_02"], 25: ["VEN_02", "ARC_01"],
        26: ["VEN_03"], 27: ["VEN_03"],
        28: ["CON_01", "CON_02", "CON_03", "CON_06"], 29: ["CON_04"],
        30: ["CON_05", "ARC_01"], 31: ["SLA_04"],
        32: ["SLA_01", "SLA_02", "ARC_01"], 33: ["SLA_03"],
        34: ["SLA_04", "ACC_01"], 35: ["SLA_04", "ACC_01"],
        36: ["SLA_04", "ACC_01"], 37: ["VEN_03", "ACC_01"],
        38: ["EXT_01", "EXT_02", "ARC_01"], 39: ["EXT_01", "EXT_02", "ARC_01"],
        40: ["EXT_01", "EXT_02", "ARC_01"], 41: ["RSK_01"],
        42: ["BCP_01", "BCP_02", "EXT_02"],
        43: ["PER_02", "PER_03", "DEL_02", "AI_03"], 44: ["RSK_03"],
        45: ["AUD_01"], 46: ["RSK_02", "REP_02"], 47: ["AUD_01"],
        48: ["REP_01"], 49: ["INC_01"], 50: ["CTX_01", "GOV_01"],
        51: ["CTX_01", "GOV_01"], 52: ["ACC_01"], 53: ["CTX_01"], 54: ["CTX_01"],
    }.items()},
    # AI guidance item mapping.
    **{("AI", f"第{i}项"): actions for i, actions in {
        1: ["GOV_01"], 2: ["AI_01", "AI_03"], 3: ["AI_01", "AI_05"], 4: ["AI_03"],
        5: ["AI_02", "AI_03"], 6: ["AI_03"], 7: ["CTX_01"], 8: ["DAT_01"], 9: ["DAT_01"],
        10: ["DAT_01"], 11: ["DEL_03"], 12: ["AI_04"], 13: ["AI_03", "AI_04"],
        14: ["RSK_01", "AI_03"], 15: ["AI_02"], 16: ["AI_02"], 17: ["AI_03", "AI_05"],
        18: ["AI_04", "CON_03", "RSK_03"], 19: ["AI_04", "RSK_03"], 20: ["AI_03"],
        21: ["AI_05", "ARC_01"], 22: ["AI_05", "AUD_01", "ARC_01"], 23: ["AI_05"],
        24: ["DAT_01", "DAT_05", "CON_03"], 25: ["AI_03", "DAT_05"], 26: ["BCP_01"],
        27: ["CTX_02"], 28: ["CTX_02"], 29: ["INC_01", "AI_02"], 30: ["CTX_02"],
        31: ["CTX_02"], 32: ["CTX_02"],
    }.items()},
}


def split_atomic_obligations(clause: base.Clause) -> list[str]:
    text = base.compact(clause.text)
    pieces = re.split(r"\n+|(?<=[。；])|(?=\d+[\.、])|(?=[（(][一二三四五六七八九十]+[）)])", text)
    pieces = [piece.strip(" \n；。") for piece in pieces if piece.strip(" \n；。")]
    modal = ("应当", "应", "不得", "须", "负责", "可以", "原则上", "鼓励", "支持", "要求")
    selected = [piece for piece in pieces if len(piece) >= 8 and any(word in piece for word in modal)]
    if not selected:
        selected = [text]
    # Keep a practical upper bound; the original full clause remains in the structured-clause sheet.
    return selected[:30]


def score_rule(text: str, rule: ActionRule) -> int:
    score = 0
    for keyword in rule.keywords:
        if keyword in text:
            score += 3 if len(keyword) >= 6 else 2
    return score


def automatic_action_match(text: str) -> list[str]:
    scores = [(score_rule(text, rule), rule.action_id) for rule in ACTION_RULES if rule.action_id != "ACT_PENDING"]
    scores = [(score, action_id) for score, action_id in scores if score > 0]
    if not scores:
        return ["ACT_PENDING"]
    scores.sort(reverse=True)
    top = scores[0][0]
    threshold = top if top < 3 else max(3, top - 2)
    return [action_id for score, action_id in scores if score >= threshold][:4]


def internal_section_fallback(clause: base.Clause, text: str) -> list[str]:
    """Classify unnumbered internal rules without pretending keyword output is final.

    These rules are intentionally applied only after the atomic-text matcher has
    failed.  The resulting status remains "待逐项抽查" in the workpaper.
    """
    if not clause.source_id.startswith("GZ_RULES:"):
        return ["ACT_PENDING"]
    source = clause.source_id
    section = clause.clause_id
    combined = f"{section} {text}"

    # Atomic-text first: some source workbooks place several logical topics in
    # one broad chapter, so the chapter label alone is not reliable enough.
    text_rules = [
        (("合同", "协议"), ["CON_01"]),
        (("风险识别", "风险评估"), ["RSK_01"]),
        (("服务水平", "服务质量", "服务目标", "考核评价"), ["SLA_04"]),
        (("退场", "离场", "资源回收", "权限回收", "工作交接"), ["EXT_02", "PER_03"]),
        (("入场", "候选人", "面试", "学历", "人员名单", "人员变更", "替换人员"), ["PER_01"]),
        (("考勤", "签到", "签退", "请假", "座位", "办公", "行员陪同"), ["PER_04"]),
        (("涉密资料", "生产数据", "内网电脑", "访问权限"), ["PER_03", "DAT_04"]),
        (("信息安全培训", "安全意识"), ["PER_02"]),
        (("自查", "检查结果", "整改措施"), ["AUD_01", "ACC_01"]),
        (("核心技术", "核心技能", "自主能力"), ["GOV_02", "DEL_03"]),
        (("集中度",), ["RSK_03"]),
        (("中止提供", "合同终止", "终止执行"), ["EXT_02"]),
        (("应急预案", "应急处置", "应急响应", "突发事件", "启动和终止"), ["INC_02"]),
        (("应急演练", "演练计划"), ["BCP_02"]),
        (("进展报告", "总结报告", "上报工作"), ["INC_01", "ARC_01"]),
        (("费用", "预算", "结算"), ["CON_01", "ARC_01"]),
        (("交付物", "产出物", "成果交接"), ["DEL_01", "DEL_03"]),
    ]
    for needles, actions in text_rules:
        if any(needle in combined for needle in needles):
            return actions

    if "专项应急预案" in source:
        if "总则" in section:
            return ["CTX_01"]
        if "应急管理" in section:
            return ["GOV_01", "INC_02"]
        return ["INC_02"]
    if "驻场服务管理细则" in source:
        if "入场" in section or "变动" in section:
            return ["PER_01"]
        if "离场" in section:
            return ["EXT_02", "PER_03"]
        return ["PER_04"]
    if "外包供应商管理细则" in source:
        if "总则" in section or "风险管理" in section:
            return ["CTX_01", "GOV_01"]
        if "选择与合同" in section:
            return ["VEN_01", "CON_01"]
        if "评价" in section:
            return ["EXT_01"]
        if "集中度" in section:
            return ["RSK_03"]
        if "退出" in section:
            return ["EXT_02"]
        if "应急" in section:
            return ["INC_02"]
        return ["SLA_04", "VEN_04"]
    if "框架协议外包" in source:
        if "总则" in section:
            return ["CTX_01", "CLS_01"]
        return ["CON_01", "SLA_04", "EXT_02"]
    if "通用外包资源使用" in source:
        return ["RSK_01", "GOV_01"]
    if "通用外包管理细则" in source:
        if "总则" in section:
            return ["CTX_01"]
        if "评价" in section:
            return ["SLA_04", "ACC_01"]
        if "安全" in section:
            return ["PER_01", "PER_04"]
        return ["CLS_01", "GOV_01"]
    if "外包管理规范" in source:
        if "总则" in section:
            return ["CTX_01"]
        if "应急" in section:
            return ["BCP_01", "INC_02"]
        return ["RSK_01", "SLA_04"]
    return ["ACT_PENDING"]


def clause_applicability(clause: base.Clause, action_ids: list[str]) -> tuple[str, str]:
    if "CTX_02" in action_ids:
        return "监管机关/行业层面条款", "不直接生成项目任务，保留为制度覆盖说明"
    if any(action_id in action_ids for action_id in ["GOV_01", "STR_01", "RSK_02", "AUD_01", "REP_02"]):
        return "机构层面适用（项目按职责配合）", "机构义务/项目配合"
    defaults = [ACTION_INDEX[action_id].default_applicability for action_id in action_ids if action_id in ACTION_INDEX]
    if any("条件" in value or "存在" in value or "发生" in value for value in defaults):
        return "条件适用，须记录触发条件和当前判断", "条件直接/支持"
    return "直接适用或作为项目制度基线", "外部监管直接或行内承接"


def build_obligations(clauses: list[base.Clause]) -> list[Obligation]:
    obligations = []
    for clause in clauses:
        override = CLAUSE_ACTION_OVERRIDES.get((clause.source_id, clause.clause_id))
        # Printing/issuance notices are provenance records, not separate project
        # duties.  Keep their clause and obligation IDs, but route them to the
        # institutional baseline instead of manufacturing operational actions.
        if clause.source_id.startswith("GZ_RULES:") and "关于印发" in clause.source_id:
            override = ["CTX_01"]
        for number, atomic_text in enumerate(split_atomic_obligations(clause), 1):
            action_ids = override or automatic_action_match(clause.title + " " + atomic_text)
            section_fallback = False
            if action_ids == ["ACT_PENDING"]:
                action_ids = internal_section_fallback(clause, atomic_text)
                section_fallback = action_ids != ["ACT_PENDING"]
            applicability, relation = clause_applicability(clause, action_ids)
            status = ("逐条规则已映射待抽查" if override else
                      "章节规则分类待逐项抽查" if section_fallback else
                      "待人工归类" if action_ids == ["ACT_PENDING"] else
                      "自动分类待抽查")
            obligations.append(Obligation(
                obligation_id=f"{clause.source_id}:{clause.clause_id}:O{number:02d}",
                source_id=clause.source_id,
                clause_id=clause.clause_id,
                clause_title=clause.title,
                atomic_text=atomic_text,
                subject=base.clause_subject(atomic_text) if len(atomic_text) > 12 else clause.subject,
                object_=base.clause_object(atomic_text),
                action_summary=atomic_text[:220],
                timing=base.clause_timing(atomic_text),
                evidence=base.clause_evidence(atomic_text),
                applicability=applicability,
                relation=relation,
                action_ids=action_ids,
                classification_status=status,
            ))
    return obligations


def load_plan_tasks() -> list[dict]:
    workbook = load_workbook(base.PLAN, data_only=True, read_only=True)
    sheet = workbook.active
    tasks = []
    for row in range(2, sheet.max_row + 1):
        tasks.append({
            "task_id": int(sheet.cell(row, 1).value),
            "stage": str(sheet.cell(row, 2).value or ""),
            "task": str(sheet.cell(row, 4).value or ""),
            "attribute": str(sheet.cell(row, 5).value or ""),
            "applicability": str(sheet.cell(row, 6).value or ""),
            "evidence": str(sheet.cell(row, 8).value or ""),
            "basis": str(sheet.cell(row, 12).value or ""),
        })
    return tasks


def map_tasks(tasks: list[dict], clauses: list[base.Clause], obligations: list[Obligation]) -> list[dict]:
    clause_index = {(clause.source_id, clause.clause_id): clause for clause in clauses}
    obligation_by_clause: dict[tuple[str, str], list[Obligation]] = {}
    for obligation in obligations:
        obligation_by_clause.setdefault((obligation.source_id, obligation.clause_id), []).append(obligation)
    output = []
    for task in tasks:
        text = " ".join([task["task"], task["attribute"], task["applicability"], task["evidence"]])
        action_ids = TASK_ACTION_OVERRIDES.get(task["task_id"], automatic_action_match(text))
        resolved, unresolved = base.resolve_basis(task["basis"], clause_index)
        cited_actions = []
        for clause in resolved:
            for obligation in obligation_by_clause.get((clause.source_id, clause.clause_id), []):
                cited_actions.extend(obligation.action_ids)
        cited_actions = list(dict.fromkeys(cited_actions))
        if task["task_id"] in TASK_ACTION_OVERRIDES:
            action_ids = list(dict.fromkeys(action_ids + cited_actions))[:6]
            method = "人工规则+制度依据复合匹配"
        elif action_ids == ["ACT_PENDING"] and cited_actions:
            action_ids = cited_actions[:5]
            method = "依据反推"
        elif cited_actions:
            # Keep text classification first, then add cited-action context without losing either route.
            action_ids = list(dict.fromkeys(action_ids + cited_actions))[:6]
            method = "任务文本+制度依据复合匹配"
        else:
            method = "任务文本匹配"
        output.append({**task, "action_ids": action_ids, "mapping_method": method,
                       "resolved_clauses": [f"{c.source_id}{c.clause_id}" for c in resolved],
                       "unresolved_basis": unresolved,
                       "mapping_status": "待人工归类" if action_ids == ["ACT_PENDING"] else "已映射待抽查"})
    return output


def build_action_catalog(obligations: list[Obligation], task_maps: list[dict]) -> list[dict]:
    catalog = []
    for action in ACTION_RULES:
        related_obligations = [item for item in obligations if action.action_id in item.action_ids]
        related_tasks = [item for item in task_maps if action.action_id in item["action_ids"]]
        clause_refs = list(dict.fromkeys(f"{item.source_id}{item.clause_id}" for item in related_obligations))
        source_ids = list(dict.fromkeys(item.source_id for item in related_obligations))
        direct_refs = list(dict.fromkeys(
            f"{item.source_id}{item.clause_id}" for item in related_obligations
            if item.relation in ["外部监管直接或行内承接", "条件直接/支持"] and item.source_id not in ["PROC"]
        ))
        catalog.append({
            **asdict(action),
            "obligation_count": len(related_obligations),
            "clause_count": len(clause_refs),
            "source_count": len(source_ids),
            "clause_refs": clause_refs,
            "obligation_ids": [item.obligation_id for item in related_obligations],
            "direct_refs": direct_refs,
            "task_ids": [item["task_id"] for item in related_tasks],
            "task_titles": [item["task"] for item in related_tasks],
            "forward_status": "待人工归类" if action.action_id == "ACT_PENDING" and related_obligations else ("有制度义务" if related_obligations else "无制度义务/项目动作"),
            "reverse_status": "已有执行计划任务" if related_tasks else "执行计划未发现对应任务",
        })
    return catalog


def build_gaps(clauses: list[base.Clause], obligations: list[Obligation], catalog: list[dict], task_maps: list[dict]) -> list[dict]:
    gaps = []
    obligation_ids_in_catalog = {
        obligation_id
        for action in catalog
        for obligation_id in action["obligation_ids"]
    }
    for obligation in obligations:
        if not obligation.action_ids:
            gaps.append({"gap_type": "正向丢失", "object_id": obligation.obligation_id, "description": "原子义务未映射到任何管理动作", "priority": "高", "suggestion": "人工分类并补充动作。"})
        elif obligation.action_ids == ["ACT_PENDING"]:
            gaps.append({"gap_type": "正向待分类", "object_id": obligation.obligation_id, "description": obligation.atomic_text[:180], "priority": "中", "suggestion": "确认管理类别和标准动作，保留原条款引用。"})
    for action in catalog:
        if action["action_id"] in ["CTX_02", "ACT_PENDING", "OPS_01"]:
            continue
        if action["obligation_count"] > 0 and not action["task_ids"]:
            gaps.append({"gap_type": "反向缺任务", "object_id": action["action_id"], "description": f"{action['title']}已有{action['clause_count']}条制度依据，但执行计划未匹配到任务", "priority": "高", "suggestion": "检查是否应新增任务、并入现有任务或记录不适用理由。"})
        if action["task_ids"] and action["obligation_count"] == 0 and action["action_type"] != "项目管理动作":
            gaps.append({"gap_type": "反向缺依据", "object_id": action["action_id"], "description": f"{action['title']}已有执行任务但未匹配制度义务", "priority": "高", "suggestion": "补充制度依据或标明为采购/合同项目动作。"})
        if action["source_count"] >= 4 and action["clause_count"] >= 12:
            gaps.append({"gap_type": "合并粒度复核", "object_id": action["action_id"], "description": f"合并后包含{action['source_count']}类来源、{action['clause_count']}条条款", "priority": "中", "suggestion": "核对动作是否过宽，必要时拆分但不得删除原义务ID。"})
    for task in task_maps:
        if task["action_ids"] == ["ACT_PENDING"]:
            gaps.append({"gap_type": "执行计划待分类", "object_id": f"TASK-{task['task_id']}", "description": task["task"], "priority": "中", "suggestion": "人工选择标准管理动作，或新增标准动作。"})

    # Classification confidence is a separate quality dimension from coverage.
    # Aggregate it by source so the gap sheet remains reviewable while the full
    # obligation-level status stays available in sheets 03 and 05.
    for status, gap_type, priority, suggestion in [
        ("逐条规则已映射待抽查", "逐条规则抽查", "中", "按制度来源抽查逐条映射；重点核对一条多义务、主体变化和条件限定，确认后再标记为已核验。"),
        ("章节规则分类待逐项抽查", "章节分类抽查", "高", "逐项回看原文；确认主体、对象、动作、时点和证据后，将结论固化为条款或原子义务规则。"),
        ("自动分类待抽查", "自动分类抽查", "中", "按管理域抽样并对高风险动作全量检查；发现错分时返回原子义务层修正规则。"),
    ]:
        grouped: dict[str, list[Obligation]] = {}
        for obligation in obligations:
            if obligation.classification_status == status:
                grouped.setdefault(obligation.source_id, []).append(obligation)
        for source_id, items in grouped.items():
            gaps.append({
                "gap_type": gap_type,
                "object_id": source_id,
                "description": f"{len(items)}项原子义务采用{status}；完整明细见03、05工作表。",
                "priority": priority,
                "suggestion": suggestion,
            })

    # Explicit merge-loss control: every obligation must remain referenced by at least one action.
    all_obligation_ids = {item.obligation_id for item in obligations}
    lost = all_obligation_ids - obligation_ids_in_catalog
    for obligation_id in sorted(lost):
        gaps.append({"gap_type": "合并引用丢失", "object_id": obligation_id, "description": "管理动作合并后未保留原子义务ID", "priority": "最高", "suggestion": "停止回写，恢复条款—义务—动作链路。"})
    return gaps


def style_sheet(sheet, freeze: str = "A2", filter_: bool = True) -> None:
    sheet.freeze_panes = freeze
    if filter_:
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def join(values: Iterable) -> str:
    return "；".join(str(value) for value in values)


def write_workbook(sources, clauses, obligations, catalog, task_maps, gaps, legacy_reviews) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "00_说明与统计"
    pending_obligations = sum(item.action_ids == ["ACT_PENDING"] for item in obligations)
    pending_tasks = sum(item["action_ids"] == ["ACT_PENDING"] for item in task_maps)
    section_review = sum(item.classification_status == "章节规则分类待逐项抽查" for item in obligations)
    automatic_review = sum(item.classification_status == "自动分类待抽查" for item in obligations)
    rule_review = sum(item.classification_status == "逐条规则已映射待抽查" for item in obligations)
    summary_data = [
        ["项目", "内容"],
        ["工作流版本", "v2.0 双向复合工作流"],
        ["正向链路", "制度源文件→结构化条款→最小可执行义务→标准管理动作"],
        ["反向链路", "标准管理动作→执行计划任务→现有制度依据→结构化条款反查"],
        ["合并控制", "标准动作合并必须保留全部原子义务ID和条款ID；丢失即阻断回写。"],
        ["制度来源", len(sources)], ["结构化条款/章节", len(clauses)], ["原子义务", len(obligations)],
        ["标准管理动作", len(ACTION_RULES) - 1], ["执行计划任务", len(task_maps)], ["双向缺口/复核提示", len(gaps)],
        ["待人工归类义务", pending_obligations], ["待人工归类任务", pending_tasks],
        ["逐条规则映射待抽查义务", rule_review],
        ["章节规则待逐项抽查义务", section_review], ["自动分类待抽查义务", automatic_review],
        ["原执行计划", "未修改"],
    ]
    for row in summary_data:
        summary.append(row)
    style_sheet(summary, filter_=False)
    set_widths(summary, [24, 110])

    source_sheet = workbook.create_sheet("01_制度源文件清单")
    source_sheet.append(["来源ID", "制度名称", "文号", "层级", "效力状态", "日期", "源文件", "格式", "结构化状态", "用途", "说明"])
    for source in sources:
        source_sheet.append([source.source_id, source.title, source.document_no, source.level, source.effect_status,
                             source.effective_date, source.path, source.format, source.extraction_status, source.review_use, source.note])
    style_sheet(source_sheet)
    set_widths(source_sheet, [14, 36, 22, 18, 20, 14, 60, 16, 34, 38, 55])

    clause_sheet = workbook.create_sheet("02_制度条款结构化")
    clause_sheet.append(["来源ID", "条款", "标题", "原文", "主体", "对象", "动作摘要", "时点频率", "证据", "校核状态"])
    for clause in clauses:
        clause_sheet.append([clause.source_id, clause.clause_id, clause.title, clause.text, clause.subject,
                             clause.object_, clause.action, clause.timing, clause.evidence, clause.extraction_status])
    style_sheet(clause_sheet)
    set_widths(clause_sheet, [24, 12, 28, 90, 34, 36, 70, 34, 45, 28])

    obligation_sheet = workbook.create_sheet("03_原子义务清单")
    obligation_sheet.append(["义务ID", "来源", "条款", "条款标题", "最小义务原文", "责任主体", "管理对象", "管理动作",
                             "时点频率", "证据", "适用性", "依据关系", "标准动作ID", "分类状态", "说明"])
    for item in obligations:
        obligation_sheet.append([item.obligation_id, item.source_id, item.clause_id, item.clause_title, item.atomic_text,
                                 item.subject, item.object_, item.action_summary, item.timing, item.evidence, item.applicability,
                                 item.relation, join(item.action_ids), item.classification_status, item.note])
    style_sheet(obligation_sheet)
    set_widths(obligation_sheet, [30, 20, 12, 24, 80, 32, 34, 65, 32, 42, 34, 28, 30, 20, 45])

    catalog_sheet = workbook.create_sheet("04_标准管理动作库")
    catalog_sheet.append(["动作ID", "管理类别", "标准管理动作", "动作类型", "默认适用性", "任务属性", "动作说明",
                          "原子义务数", "条款数", "来源数", "保留的全部义务ID", "全部条款依据", "直接/条件依据",
                          "执行计划序号", "正向状态", "反向状态"])
    for item in catalog:
        catalog_sheet.append([item["action_id"], item["category"], item["title"], item["action_type"],
                              item["default_applicability"], item["task_attribute"], item["description"],
                              item["obligation_count"], item["clause_count"], item["source_count"], join(item["obligation_ids"]), join(item["clause_refs"]),
                              join(item["direct_refs"]), join(item["task_ids"]), item["forward_status"], item["reverse_status"]])
    style_sheet(catalog_sheet)
    set_widths(catalog_sheet, [14, 24, 42, 22, 34, 22, 60, 12, 10, 10, 90, 75, 65, 35, 22, 24])

    forward_sheet = workbook.create_sheet("05_条款到动作正向矩阵")
    forward_sheet.append(["来源", "条款", "义务ID", "最小义务", "适用性", "标准动作ID", "标准动作名称", "分类状态", "原条款是否保留"])
    for item in obligations:
        for action_id in item.action_ids:
            action = ACTION_INDEX[action_id]
            forward_sheet.append([item.source_id, item.clause_id, item.obligation_id, item.atomic_text, item.applicability,
                                  action_id, action.title, item.classification_status, "是"])
    style_sheet(forward_sheet)
    set_widths(forward_sheet, [22, 12, 30, 85, 34, 16, 45, 20, 18])

    reverse_sheet = workbook.create_sheet("06_动作到制度反向矩阵")
    reverse_sheet.append(["动作ID", "管理类别", "标准动作", "保留的全部义务ID", "直接/条件制度依据", "全部制度依据", "对应执行计划序号",
                          "对应执行事项", "无任务时处理", "合并后条款引用是否保留"])
    for item in catalog:
        if item["action_id"] == "ACT_PENDING":
            continue
        reverse_sheet.append([item["action_id"], item["category"], item["title"], join(item["obligation_ids"]), join(item["direct_refs"]),
                              join(item["clause_refs"]), join(item["task_ids"]), "\n".join(item["task_titles"]),
                              "已有任务" if item["task_ids"] else "需新增/并入/记录不适用理由", "是"])
    style_sheet(reverse_sheet)
    set_widths(reverse_sheet, [14, 24, 42, 90, 70, 80, 34, 65, 38, 22])

    task_sheet = workbook.create_sheet("07_执行计划到动作映射")
    task_sheet.append(["序号", "阶段", "执行事项", "任务属性", "适用性", "当前制度依据", "标准动作ID", "标准动作名称",
                       "匹配方式", "已解析条款", "未解析依据", "映射状态"])
    for task in task_maps:
        task_sheet.append([task["task_id"], task["stage"], task["task"], task["attribute"], task["applicability"],
                           task["basis"], join(task["action_ids"]), join(ACTION_INDEX[i].title for i in task["action_ids"]),
                           task["mapping_method"], join(task["resolved_clauses"]), join(task["unresolved_basis"]), task["mapping_status"]])
    style_sheet(task_sheet)
    set_widths(task_sheet, [8, 20, 44, 20, 48, 65, 34, 70, 26, 45, 50, 20])

    gap_sheet = workbook.create_sheet("08_双向缺口清单")
    gap_sheet.append(["缺口类型", "对象ID", "说明", "优先级", "建议处理", "人工确认结论", "确认人/日期"])
    for gap in gaps:
        gap_sheet.append([gap["gap_type"], gap["object_id"], gap["description"], gap["priority"], gap["suggestion"], "待确认", ""])
    style_sheet(gap_sheet)
    set_widths(gap_sheet, [22, 30, 80, 12, 65, 28, 22])

    legacy_sheet = workbook.create_sheet("09_原制度依据问题")
    legacy_sheet.append(["序号", "执行事项", "当前制度依据", "复核结论", "问题类型", "建议处置", "建议直接依据", "建议支持依据"])
    for item in legacy_reviews:
        if item["conclusion"] in ["初审未发现明显错引", "项目依据可保留"]:
            continue
        legacy_sheet.append([item["task_id"], item["task"], item["current_basis"], item["conclusion"], item["issue_type"],
                             item["advice"], item["suggested_direct"], item["suggested_support"]])
    style_sheet(legacy_sheet)
    set_widths(legacy_sheet, [8, 44, 70, 28, 34, 70, 55, 55])

    confirmation_sheet = workbook.create_sheet("10_人工确认与回写")
    confirmation_sheet.append(["对象类型", "对象ID", "当前结论", "最终分类/动作", "最终直接依据", "最终支持依据",
                               "不适用理由/触发条件", "是否允许回写", "确认人/日期", "备注"])
    for gap in gaps:
        confirmation_sheet.append([gap["gap_type"], gap["object_id"], "待确认", "", "", "", "", "否", "", ""])
    style_sheet(confirmation_sheet)
    set_widths(confirmation_sheet, [22, 30, 22, 45, 65, 65, 55, 18, 22, 45])

    rule_sheet = workbook.create_sheet("11_双向规则与反馈")
    rule_sheet.append(["规则", "内容", "阻断条件"])
    rules = [
        ("B1 正向完整性", "每条制度条文拆成最小义务，每个适用义务至少映射一个标准管理动作。", "存在未映射义务。"),
        ("B2 反向有据性", "每个监管型标准动作至少有一项直接、条件或行内承接依据。", "监管型动作无制度依据。"),
        ("B3 执行覆盖", "每个适用于本项目的标准动作应有执行计划任务，或明确并入/不适用理由。", "适用动作无任务且无解释。"),
        ("B4 合并不丢失", "管理动作分类合并必须保留全部原子义务ID和条款ID的并集。", "出现合并引用丢失。"),
        ("B5 层级分离", "监管直接、行内承接、采购合同直接和支持依据分列。", "以支持依据替代直接依据。"),
        ("B6 条件可追溯", "条件适用动作记录触发条件、当前判断和重新判断时点。", "只写条件适用但无触发事实。"),
        ("B7 机构项目分离", "机构治理义务和项目配合动作分别表述。", "将董事会、高管或审计职责直接分派项目组。"),
        ("B8 人工回写门", "双向缺口和原文校核关闭后方可修改执行计划。", "人工确认列未允许回写。"),
    ]
    for row in rules:
        rule_sheet.append(row)
    style_sheet(rule_sheet)
    set_widths(rule_sheet, [20, 90, 55])

    workbook.save(OUTPUT)


def write_json(obligations, catalog, task_maps, gaps) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    (DATA_DIR / "原子义务结构化.json").write_text(json.dumps([asdict(item) for item in obligations], ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "标准管理动作库.json").write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "执行计划到标准动作映射.json").write_text(json.dumps(task_maps, ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "双向覆盖缺口.json").write_text(json.dumps(gaps, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    clauses = base.build_clauses()
    legacy_reviews = base.review_plan(clauses)
    obligations = build_obligations(clauses)
    tasks = load_plan_tasks()
    task_maps = map_tasks(tasks, clauses, obligations)
    catalog = build_action_catalog(obligations, task_maps)
    gaps = build_gaps(clauses, obligations, catalog, task_maps)
    write_workbook(base.SOURCE_SPECS, clauses, obligations, catalog, task_maps, gaps, legacy_reviews)
    write_json(obligations, catalog, task_maps, gaps)
    print(json.dumps({
        "output": str(OUTPUT), "sources": len(base.SOURCE_SPECS), "clauses": len(clauses),
        "obligations": len(obligations), "standard_actions": len(ACTION_RULES) - 1,
        "tasks": len(task_maps), "gaps": len(gaps),
        "pending_obligations": sum(item.action_ids == ["ACT_PENDING"] for item in obligations),
        "pending_tasks": sum(item["action_ids"] == ["ACT_PENDING"] for item in task_maps),
    }, ensure_ascii=False))


if __name__ == "__main__":
    main()
