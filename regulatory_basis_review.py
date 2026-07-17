#!/usr/bin/env python3
"""Generate a reusable regulatory-basis review workbook for the outsourcing plan.

The script deliberately does not edit the source execution plan.  It performs:

1. source inventory;
2. clause extraction and normalization;
3. citation resolution;
4. seven-question task review;
5. issue-list generation;
6. machine-readable JSON export.

It is designed for iterative review: manual decisions can later be added to
REVIEW_OVERRIDES without changing the source workbook.
"""

from __future__ import annotations

import argparse
import json
import re
import zlib
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
PLAN = ROOT / "外包商管理全流程执行计划-授信系统群咨询项目组.xlsx"
REG_DIR = ROOT / "银行外包风险管理监管要求"
INTERNAL_3 = ROOT / "管理办法：信息科技外包风险管理办法（3.0版，2026年）.xlsx"
INTERNAL_2 = ROOT / "管理办法：信息科技外包风险管理办法（2.0版，2024年）.xlsx"
INTERNAL_RULES = ROOT / "管理办法：信息科技外包供应商管理细则（1.0版，2025年）》等七份制度.xlsx"
OUTPUT = ROOT / "制度依据专项复核工作底稿-授信系统群咨询项目.xlsx"
DATA_DIR = ROOT / "制度依据专项复核数据"


CN_DIGITS = {"零": 0, "〇": 0, "一": 1, "二": 2, "两": 2, "三": 3, "四": 4,
             "五": 5, "六": 6, "七": 7, "八": 8, "九": 9}
CN_UNITS = {"十": 10, "百": 100, "千": 1000}


def cn_to_int(value: str) -> int:
    if value.isdigit():
        return int(value)
    total = 0
    current = 0
    for ch in value:
        if ch in CN_DIGITS:
            current = CN_DIGITS[ch]
        elif ch in CN_UNITS:
            unit = CN_UNITS[ch]
            if current == 0:
                current = 1
            total += current * unit
            current = 0
    return total + current


def int_to_cn(num: int) -> str:
    digits = "零一二三四五六七八九"
    if num < 10:
        return digits[num]
    if num < 20:
        return "十" + (digits[num % 10] if num % 10 else "")
    if num < 100:
        return digits[num // 10] + "十" + (digits[num % 10] if num % 10 else "")
    return str(num)


def compact(text: str) -> str:
    text = text.replace("\u3000", " ").replace("\xa0", " ").replace("\u2002", " ").replace("\u2003", " ")
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def compact_pdf_text(text: str) -> str:
    # PDF glyph extraction often produces one Chinese character per text run.
    text = text.replace("�", "")
    return re.sub(r"\s+", "", text)


@dataclass
class Source:
    source_id: str
    title: str
    document_no: str
    level: str
    effect_status: str
    effective_date: str
    path: str
    format: str
    extraction_status: str
    review_use: str
    note: str = ""


@dataclass
class Clause:
    source_id: str
    clause_id: str
    title: str
    text: str
    subject: str
    object_: str
    action: str
    timing: str
    evidence: str
    extraction_status: str = "自动结构化，待人工校核"


SOURCE_SPECS = [
    Source("2009", "商业银行信息科技风险管理指引", "银监发〔2009〕19号", "外部监管", "现行基础性指引", "2009", str(REG_DIR / "《商业银行信息科技风险管理指引》（银监发[2009]19号印发）.pdf"), "PDF", "轮廓化PDF，无法可靠自动抽取条款", "直接依据/补充依据，逐条人工校核", "已结构化第32、45、62条；其余条款不得仅凭编号自动判定。"),
    Source("141", "银行保险机构信息科技外包风险监管办法", "银保监办发〔2021〕141号", "外部监管", "现行有效", "2021-12-30", str(REG_DIR / "银行保险机构信息科技外包风险监管办法_银保监办发2021_141号.md"), "Markdown", "可自动结构化", "信息科技外包主干直接依据"),
    Source("DATA", "银行保险机构数据安全管理办法", "金规〔2024〕24号", "外部监管", "现行有效", "2024-12-27", str(REG_DIR / "国家金融监督管理总局关于印发银行保险机构数据安全管理办法的通知.pdf"), "PDF", "可通过ToUnicode映射结构化，需抽样校核", "数据处理场景直接或条件依据"),
    Source("AI", "银行业保险业人工智能安全开发应用指导意见", "金发〔2026〕8号", "外部监管", "现行指导意见", "2026-06-18", str(REG_DIR / "银行业保险业人工智能安全开发应用指导意见_金发2026_8号.md"), "Markdown", "可自动结构化", "AI实际使用、模型及方案设计条件依据"),
    Source("NET_DRAFT", "银行业保险业网络安全管理办法（征求意见稿）", "征求意见稿", "外部监管趋势", "未生效", "2026-07-10", str(REG_DIR / "国家金融监督管理总局关于《银行业保险业网络安全管理办法（征求意见稿）》.pdf"), "PDF", "可抽取但不作为正式依据", "仅作前瞻参考，不得列作现行直接依据"),
    Source("GZ3", "广州银行信息科技外包风险管理办法（3.0版，2026年）", "行内制度", "行内制度", "现行有效", "2026", str(INTERNAL_3), "XLSX提取汇总", "可自动结构化", "行内承接直接依据"),
    Source("GZ2", "广州银行信息科技外包风险管理办法（2.0版，2024年）", "行内制度", "行内制度", "已被3.0版替代", "2024", str(INTERNAL_2), "XLSX提取汇总", "可读取，仅用于版本沿革", "不得作为当前直接依据"),
    Source("GZ_RULES", "广州银行信息科技外包供应商、驻场、通用外包等七份配套制度", "行内配套制度", "行内制度", "按现有资料视为现行，需制度归口确认", "2025", str(INTERNAL_RULES), "XLSX提取汇总", "部分文件可结构化，部分ZIP行跳过但另有解析副本", "供应商、驻场、应急、框架及资源管理依据"),
    Source("PROC", "授信系统群咨询项目采购文件及交付物要求", "项目采购依据", "项目管理依据", "本项目有效", "本项目", str(ROOT / "授信系统群咨询项目采购交付物与外包管理执行计划支撑分析.md"), "Markdown/采购材料", "已有交付物支撑分析", "交付范围、验收、付款及服务边界的直接项目依据"),
]


SOURCE_ALIASES = {
    "141": ["银行保险机构信息科技外包风险监管办法", "141号文"],
    "2009": ["商业银行信息科技风险管理指引", "2009年指引"],
    "DATA": ["银行保险机构数据安全管理办法", "数据安全办法"],
    "AI": ["银行业保险业人工智能安全开发应用指导意见", "人工智能安全开发应用指导意见", "AI指导意见"],
    "GZ3": ["广州银行信息科技外包风险管理办法（3.0", "广州银行信息科技外包风险管理办法(3.0"],
    "PROC": ["采购文件", "采购交付", "采购内容", "付款条件", "合同"],
}


INTERNAL_SECTION_HEADINGS = [
    "总则", "职责与分工", "外包驻场分类与定义", "外包商分类分级管理", "外包商准入与尽职调查",
    "外包商选择与合同管理", "外包商评价管理", "外包商风险管理", "外包商退出管理",
    "尽职调查流程", "尽职调查结果", "尽职调查结果报送", "尽职调查有效期",
    "入场管理", "行为规范管理", "考勤管理", "变动管理", "离场管理", "处罚管理",
    "外包商服务退出监测", "外包商退出业务影响分析", "外包商退出交接安排",
    "集中度管理要求", "风险监控管理要求", "评价管理", "合同管理", "应急管理",
]


REVIEW_PASS_OVERRIDES = {
    1: "制度基线盘点属于方法性任务，列示全部现行制度及效力状态是适当依据表达，不要求绑定单一条款。",
    2: "覆盖矩阵属于制度落实方法，行内第50条与外部适用条款清单共同作为依据可以保留。",
    15: "行内3.0版第12至14条与供应商管理细则职责分工共同覆盖相关部室、执行部门和项目团队职责，组合引用与任务一致。",
    19: "本项任务就是判断三类配套制度的适用范围，引用各制度适用范围属于直接依据；当前不适用判断需随合作模式变化复核。",
    22: "数据安全评估、委托处理、共同处理等模式判断分别对应数据安全办法第22、30至32条，条款组合与任务覆盖范围一致。",
    23: "个人信息保护影响评估及委托处理保护要求对应数据安全办法第58、61至63条，属于条件触发下的直接依据组合。",
    40: "一级外包商强化管理要求在供应商管理细则中有明确规定，当前按分类分级结论条件适用的表达正确。",
    57: "数据委托、共同处理及个人信息处理的合同边界分别由数据安全办法第30至32、61条支撑；组合引用与任务覆盖的数据处理模式一致。",
    58: "数据委托、共同处理及个人信息委托处理的协议要求分别对应第30至32、61条，在实际处理模式触发后直接适用。",
    108: "任务同时覆盖AI透明度、可解释性、公平伦理和审计追溯，AI指导意见第21至23项与任务对象和动作一致，可整体保留。",
    130: "定期复核制度效力属于方法性持续任务，列示全部监管和行内制度基线是适当表达。",
    131: "定期维护覆盖矩阵属于制度落实方法，外部适用条款作为复核范围而非某项直接义务，可保留但应避免被理解为具体任务依据。",
}


# Confirmed/manual semantics take precedence over heuristic extraction.
MANUAL_CLAUSES = [
    Clause("2009", "第32条", "重大信息科技项目管理", "商业银行应有能力对信息系统进行需求分析、采购、开发、测试、部署、维护、升级和报废，制定制度和流程，管理信息科技项目的优先排序、立项、审批和控制。项目实施部门应定期向信息科技管理委员会提交重大信息科技项目的进度报告，由其进行审核，进度报告应当包括计划的重大变更、关键人员或供应商的变更以及主要费用支出情况。应在信息系统投产后一定时期内组织后评价并优化。", "商业银行；项目实施部门；信息科技管理委员会", "重大信息科技项目及投产后的信息系统", "建立项目管理流程；定期提交并审核进度报告；投产后组织后评价和优化", "重大项目实施期间定期；系统投产后一定时期", "重大项目进度报告、信科委审核记录、后评价报告及优化记录", "已由行内风险评估表引用原文校核"),
    Clause("2009", "第45条", "信息科技运行服务水平管理", "商业银行应建立服务水平管理相关的制度和流程，对信息科技运行服务水平进行考核。", "商业银行", "信息科技运行服务水平", "建立运行服务水平管理制度和流程并实施考核", "按制度规定持续/周期开展", "运行服务水平制度、流程和考核记录", "用户提供原文，已校核"),
    Clause("2009", "第62条", "外包合同审核及SLA审阅修订", "商业银行所有信息科技外包合同应由信息科技风险管理部门、法律部门和信息科技管理委员会审核通过。商业银行应设立流程定期审阅和修订服务水平协议。", "商业银行；信息科技风险管理部门；法律部门；信息科技管理委员会", "所有信息科技外包合同及服务水平协议", "三方审核外包合同；建立流程定期审阅并修订SLA", "合同签署/变更前；SLA按流程定期审阅并在触发时修订", "三方审核意见和结论、SLA审阅记录、修订版本及审批记录", "用户提供原文，已校核"),
]


MANUAL_141 = {
    15: ("银行保险机构", "拟开展的信息科技外包活动", "评估战略一致性和相关风险并审慎决策；重要外包至少向高管层报告审批", "外包实施决策前", "风险评估、决策及重要外包高管层报告审批记录"),
    21: ("银行保险机构", "信息科技外包合同或协议", "在合同中明确服务、风险、安全、检查审计、变更终止、报告等事项", "合同签订前；变更时复核", "合同条款、法审和风险审核记录"),
    23: ("银行保险机构", "外包服务过程", "持续监控并及时发现、纠正异常", "服务期间持续", "监控记录、异常台账及整改关闭记录"),
    24: ("银行保险机构", "信息科技外包服务目录、SLA及监控评价", "建立服务目录、SLA和监控评价机制，确保数据真实完整并保存", "服务期间持续；数据至少保存至服务结束后三年", "服务目录、SLA、监控评价数据及归档记录"),
    25: ("银行保险机构", "信息科技外包服务效能和质量", "建立服务效能和质量监控指标并实施监控", "服务期间按指标频率持续/周期开展", "指标定义、监控结果、评价和整改记录"),
    29: ("银行保险机构", "到期或结束的外包服务及服务提供商", "到期前评估是否继续外包；结束时评价；持续性服务退出前制定退出交接计划", "到期前；服务结束时；终止或更换前", "续包决策、服务商评价、退出和交接计划"),
    30: ("银行保险机构", "信息科技外包风险", "建立完善风险管理制度流程，充分识别评估科技能力、业务中断、数据、资金、服务水平等风险", "外包前并在变化或持续管理中更新", "风险制度、风险识别评估及控制措施记录"),
    31: ("银行保险机构", "可能重大影响业务连续性的重要外包服务及服务提供商", "建立风险控制缓释转移措施，组织服务商参与应急计划和演练并保留最低接管能力", "事先建立；实施期间持续；至少每年演练纳入一个或多个服务商", "退出和供应链方案、应急预案、年度演练及整改记录"),
    32: ("银行保险机构", "服务提供商、外包人员及外包活动中的信息资产、交付物、敏感信息、模型算法", "落实培训、保密、最小授权、安全检查、敏感信息监测、模型算法管理及定期安全评估", "入场/授权/交付前及服务期间持续；安全评估定期", "培训保密、权限、安全扫描、敏感信息监测、模型管理和评估记录"),
    33: ("银行保险机构", "具有集中度风险的外包服务及服务提供商", "识别集中度风险并通过分散、知识产权、自身能力和替代商等措施降低依赖", "选型时及服务期间持续/周期复核", "集中度评估、替代方案和风险缓释记录"),
    35: ("银行保险机构", "信息科技外包风险管理整体情况", "开展全面风险管理评估并向董事会或高级管理层报告", "每年至少一次", "年度评估报告及治理层报送/审议记录"),
    36: ("银行保险机构；内部审计", "信息科技外包及其风险管理", "开展定期审计，三年覆盖重要外包；重大事件后专项审计", "定期；至少每三年覆盖全部重要外包；重大事件后及时", "审计计划、报告、整改及专项审计记录"),
    37: ("银行保险机构", "列明的重大/重要信息科技外包活动", "向监管部门履行事前报告", "外包合同签订前二十个工作日", "监管报告材料、报送和反馈记录"),
    38: ("银行保险机构", "信息科技外包重大风险事件", "按突发事件监管要求报告；无明确规定时在24小时内报告", "事件发生后按规定；兜底24小时内", "事件报告、报送回执、处置和复盘记录"),
}


# Confirmed task-level conclusions.  These are review opinions, not workbook edits.
REVIEW_OVERRIDES = {
    21: ("直接与条件依据混杂", "保留141号文第15、30条为综合外包风险评估直接依据；2009年第32条仅在重大项目正式触发时使用；数据安全条款仅在数据处理场景触发时使用。", "141号文第15、30条", "2009年第32条、数据安全办法第22条按条件作为补充/专项依据", "拆分综合评估主依据与重大项目、数据、AI专项模块。"),
    50: ("2009年条款原文待核但存在现行直接依据", "保留141号文第21条第4项作为检查、评估和审计权的直接依据；2009年第58条第2项回原文确认后再决定并列层级。", "141号文第21条第4项", "2009年第58条第2项待原文核验", "核对历史指引条款原文，不影响141号文直接依据成立。"),
    51: ("2009年条款原文待核但存在现行直接依据", "保留141号文第21条第5项作为合同变更、终止和过渡安排直接依据；2009年第57条、第58条第8项核原文后再决定保留。", "141号文第21条第5项", "2009年第57条、第58条第8项待核", "“重新评估”与“合同变更终止条款”应分清直接来源。"),
    62: ("直接依据与补充依据混杂", "保留采购文件和141号文第21条为服务边界直接依据；第24条用于服务目录/SLA，第29条用于到期及退出，不宜整体并列为服务边界直接依据。", "采购文件；141号文第21条", "141号文第24、29条", "按合同边界、服务目录、到期退出三个动作拆分。"),
    71: ("监管条款被扩张用于项目管理动作", "采购交付清单是D01—D42基线的直接依据；141号文第24、25条降为后续服务监控和质量指标的支持依据。", "采购文件交付内容清单", "141号文第24、25条", "不得由SLA监控条款直接推出具体交付物基线。"),
    84: ("2009年条款原文待核", "保留行内驻场人员入场管理要求；2009年第40条在未核原文前不得作为直接依据。", "行内驻场服务管理细则人员入场要求", "2009年第40条待原文核验", "重点核对第40条是否确指外包人员资质和背景。"),
    88: ("培训义务与培训内容依据混杂", "2009年第31条和数据安全办法第15条核实后作为培训义务依据；AI指导意见第24、25项用于确定AI工具使用培训内容，不直接要求本项目开展入场培训。", "数据安全办法第15条；141号文第32条第1项可补充外包人员安全培训", "AI指导意见第24、25项作为培训内容支持；2009年第31条待核", "区分‘必须培训’与‘培训讲什么’。"),
    92: ("直接外部依据遗漏且历史条款混用", "补充141号文第32条第2项作为最小授权直接依据；2009年第17、22、60条逐条核实后再决定保留层级。", "141号文第32条第2项；数据安全办法第28、43条", "2009年第17、22、60条待核", "权限配置与定期复核应分别核对。"),
    97: ("数据安全闸门引用包含条件不一致条款", "第22、28、30至32条可分别支持评估、授权、委托和共同处理判断；第29条主要针对集团内数据共享，不应因连续范围引用自动适用于本项目。", "数据安全办法第22、28、30至32条按实际模式", "第29条仅发生对应共享场景时适用", "将每批资料查看、数据提取、委托处理、共同处理分别设置判断分支。"),
    98: ("质量监控条款不能直接推出阶段评审", "采购文件、合同及验收标准是阶段评审直接依据；141号文第24、25条仅支持质量监控和评价。", "采购文件、合同和验收标准", "141号文第24、25条", "阶段评审与监管SLA监控不可等同。"),
    101: ("监管实施义务被转化为咨询交付要求", "数据安全办法第42至48条直接约束银行的数据和系统保护；本咨询任务是把要求写入蓝图，属于项目承接措施，应将采购交付要求列为直接依据、监管条款列为内容依据。", "采购交付要求、合同及高阶需求评审标准", "数据安全办法第42至48条作为蓝图内容依据", "后续建设投产阶段再直接落实技术控制。"),
    102: ("交付物检查动作与监管控制依据混杂", "采购/合同的交付评审要求是检查动作直接依据；141号文第32条和数据安全办法相关条款用于确定检查内容，并按是否涉及敏感数据和传输条件适用。", "采购文件、合同和交付评审标准", "141号文第32条；数据安全办法第28、42至44条按场景", "第44条仅在发生数据传输时直接相关。"),
    104: ("AI引用范围过宽且评审动作需分层", "数据安全办法第50至52条可作为模型算法审查内容依据；AI指导意见第14至26项范围过宽，应按模型治理、分类分级、高风险准入、外包、供应链、透明可解释等具体成果拆分。", "采购交付要求和模型/算法方案评审标准", "数据安全办法第50至52条；AI指导意见按具体项拆分", "方案评审不等同于模型已投入使用，需区分咨询阶段与后续建设运行阶段。"),
    114: ("重大项目条款与普通阶段评审混用", "采购文件是D11—D30评审直接依据；2009年第32条仅在重大项目正式认定后支持信科委报告审核；数据和AI条款应按具体成果内容精确到条/项。", "采购文件、合同和验收标准", "2009年第32条、数据安全及AI具体条款按条件", "“AI指导意见”整份文件引用过宽。"),
    116: ("质量监控条款不能直接推出阶段评审", "采购文件、合同及验收标准作为直接依据；141号文第24、25条降为质量监控支持依据。", "采购文件、合同和验收标准", "141号文第24、25条", "需区分交付评审与外包质量监控。"),
    117: ("付款门的依据层级混杂", "采购文件、正式合同和付款条件是付款合规门直接依据；141号文第24、25条仅支持服务评价，不直接设定付款门。", "采购文件、正式合同及付款条件", "141号文第24、25条", "监管条款不替代合同付款约定。"),
    118: ("验收依据与监管支持依据混杂", "采购文件、合同和逐项验收标准是直接依据；141号文第24、25条降为质量监控支持依据。", "采购文件、合同和验收标准", "141号文第24、25条", "逐项验收标准必须来自采购/合同。"),
    124: ("条款范围过宽", "按持续监控、质量指标、异常整改、到期评价和退出分别引用第23、25、27、29条；第28条仅在关联外包时适用。", "采购文件；141号文第23、25、27条", "第24、26、29条按具体动作；第28条仅关联外包", "不得以第23至29条整体覆盖全部持续服务动作。"),
    128: ("2009年条款待原文核验", "以数据安全办法返还、删除、销毁及不可恢复要求为直接依据；2009年第60条第6项核对原文后决定是否保留。", "数据安全办法第38、46条", "2009年第60条第6项待核", "分别核对资料、数据、介质和实物的处置要求。"),
    129: ("任务属于后续移交，直接监管依据需分层", "内部审计义务可补充141号文第36条；本项‘移交’本身是项目衔接措施，不应表述为第66条直接要求的当前项目动作。", "项目收尾及后续建设衔接安排", "141号文第36条；2009年第66条待原文核验", "明确后续建设触发条件和承接主体。"),
    136: ("机构治理依据与项目配合动作需分层", "141号文第7、8条及2009年治理条款直接要求机构治理主体履职；本项目仅承担材料准备和配合，应保留机构层面适用表述。", "机构治理制度和会议安排", "141号文第7、8条；2009年第7、32条按原文和重大项目条件", "不能将董事会、高管层义务改写为项目团队直接责任。"),
    148: ("培训触发与培训内容依据混杂", "培训义务依据与AI、数据安全控制内容分开；新增人员、制度变化和关键阶段的频率主要来自行内培训机制。", "行内培训制度；数据安全办法第15条；141号文第32条第1项", "AI指导意见第24、25项作为内容支持；2009年第31条待核", "外部条款未必直接规定每次关键阶段都要补训。"),
    149: ("直接外部依据遗漏且历史条款混用", "补充141号文第32条第2项支持访问授权持续管控；2009年第17、22、60条核原文后分直接与支持。", "141号文第32条第2项；数据安全办法第28、43条", "2009年第17、22、60条待核", "区分初始授权、周期复核和离场注销。"),
    160: ("服务目录条款被扩张用于交付台账", "采购交付清单是交付物计划、责任和版本台账直接依据；141号文第24条仅在该台账同时承担服务目录/监控功能时作为支持依据。", "采购交付清单、合同和项目计划", "141号文第24条", "标明交付物台账是否兼具服务目录属性。"),
    170: ("直接外部依据遗漏", "补充141号文第31条第4项作为重要外包服务商参与应急计划和演练的直接依据；2009年第50至54条保留为机构业务连续性支持依据。", "141号文第31条第4项（满足重要且影响连续性条件时）", "2009年第50至54条；数据安全办法事件条款", "本项目是否触发取决于重要外包及业务连续性影响判断。"),
    171: ("机构层面要求与项目配合动作混用", "业务连续性计划和年度演练是机构层面事项；如本项目满足141号文第31条条件，再要求服务商提供并参与。", "机构业务连续性制度；141号文第31条按条件", "2009年第54条", "不能无条件要求普通咨询项目单独每年演练。"),
    173: ("外部监管直接依据遗漏", "补充141号文第33条作为识别和降低外包集中度风险的直接依据。", "141号文第33条；行内3.0版第44条", "供应商管理细则集中度要求", "单一中标商不自动等于高集中度，但必须识别评估。"),
    185: ("保存义务范围被扩张", "141号文第24条仅直接支持SLA监控信息和评价结果保存；数据日志和AI记录应分别适用数据安全、AI条款；一般项目档案按行内档案及采购合同要求。", "各证据类型对应的专项条款和行内档案制度", "141号文第24条仅限SLA监控评价证据", "应建立证据类型—保存期限—依据的子矩阵。"),
    190: ("直接与条件依据混杂", "141号文第15、30条作为重大变化后重评估的主依据；重大项目、数据和AI条款按变化内容条件适用。", "141号文第15、30条", "2009年第32、57条及数据安全第22条按条件", "重新评估触发条件需与各专项制度分别对应。"),
    193: ("两个触发动作需拆分", "风险变化触发重新评估；形成合同或协议重大变更时，再适用第62条三方合同审核。", "2009年第57条（核原文后）；合同变更时第62条", "行内合同变更制度", "仅项目安排变化但未形成合同变更时，不自动触发三方合同审核。"),
}


def extract_pdf_objects(path: Path) -> dict[int, bytes]:
    data = path.read_bytes()
    return {int(m.group(1)): m.group(2) for m in re.finditer(rb"(\d+)\s+0\s+obj\s*(.*?)\s*endobj", data, re.S)}


def pdf_stream(objects: dict[int, bytes], number: int) -> bytes:
    obj = objects[number]
    match = re.search(rb"stream\r?\n(.*?)endstream", obj, re.S)
    if not match:
        return b""
    value = match.group(1)
    if b"/FlateDecode" in obj[:match.start()]:
        value = zlib.decompress(value)
    return value.rstrip(b"\r\n")


def parse_cmap(value: bytes) -> dict[int, str]:
    result: dict[int, str] = {}
    for match in re.finditer(rb"<([0-9A-Fa-f]{4})>\s*<([0-9A-Fa-f]{4})>\s*\[(.*?)\]", value, re.S):
        start, end = int(match.group(1), 16), int(match.group(2), 16)
        targets = re.findall(rb"<([0-9A-Fa-f]+)>", match.group(3))
        for offset, target in enumerate(targets[:end - start + 1]):
            try:
                result[start + offset] = bytes.fromhex(target.decode()).decode("utf-16-be")
            except (ValueError, UnicodeDecodeError):
                pass
    for source, target in re.findall(rb"<([0-9A-Fa-f]{4})>\s*<([0-9A-Fa-f]+)>", value):
        try:
            result[int(source, 16)] = bytes.fromhex(target.decode()).decode("utf-16-be")
        except (ValueError, UnicodeDecodeError):
            pass
    return result


def extract_pdf_text(path: Path) -> str:
    objects = extract_pdf_objects(path)
    font_maps: dict[int, dict[int, str]] = {}
    for number, obj in objects.items():
        if b"/Subtype /Type0" not in obj:
            continue
        match = re.search(rb"/ToUnicode\s+(\d+)\s+0\s+R", obj)
        if match:
            font_maps[number] = parse_cmap(pdf_stream(objects, int(match.group(1))))

    pages: list[tuple[int, bytes, dict[str, int]]] = []
    for number, obj in objects.items():
        if not re.search(rb"/Type\s*/Page(?!s)\b", obj):
            continue
        content_match = re.search(rb"/Contents\s+(\d+)\s+0\s+R", obj)
        if not content_match:
            continue
        resources = obj
        resource_match = re.search(rb"/Resources\s+(\d+)\s+0\s+R", obj)
        if resource_match:
            resources = objects.get(int(resource_match.group(1)), b"")
        fonts = {name.decode(): int(ref) for name, ref in re.findall(rb"/([A-Za-z]+\d+)\s+(\d+)\s+0\s+R", resources)}
        pages.append((number, pdf_stream(objects, int(content_match.group(1))), fonts))

    token = re.compile(
        rb"/([A-Za-z]+\d+)\s+[\d.]+\s+Tf|<([0-9A-Fa-f]+)>\s*Tj|\[(.*?)\]\s*TJ|T\*|[-\d.]+\s+[-\d.]+\s+Td",
        re.S,
    )
    output: list[str] = []
    for _, content, fonts in pages:
        font_number = None
        for match in token.finditer(content):
            if match.group(1):
                font_number = fonts.get(match.group(1).decode())
                continue
            if match.group(2):
                hex_value = match.group(2)
                cmap = font_maps.get(font_number or -1, {})
                output.append("".join(cmap.get(int(hex_value[i:i + 4], 16), "") for i in range(0, len(hex_value), 4)))
            elif match.group(3):
                cmap = font_maps.get(font_number or -1, {})
                for hex_value in re.findall(rb"<([0-9A-Fa-f]+)>", match.group(3)):
                    output.append("".join(cmap.get(int(hex_value[i:i + 4], 16), "") for i in range(0, len(hex_value), 4)))
            else:
                output.append("\n")
        output.append("\n")
    return "".join(output)


def clause_subject(text: str) -> str:
    candidates = [
        "银行保险机构", "商业银行", "信息科技外包活动执行部门", "信息科技风险管理部门",
        "项目实施部门", "信息科技管理委员会", "董事会", "高级管理层", "金融机构",
        "数据安全归口管理部门", "信息科技部门", "服务提供商", "外包人员",
    ]
    found = [item for item in candidates if item in text[:180]]
    return "；".join(found) if found else "需结合完整条款和行内职责确认"


def clause_object(text: str) -> str:
    groups = [
        ("外包合同/协议", ["外包合同", "合同或协议", "外包协议"]),
        ("服务提供商/外包商", ["服务提供商", "外包商"]),
        ("外包人员/驻场人员", ["外包人员", "驻场人员"]),
        ("信息科技外包活动", ["信息科技外包活动", "外包服务过程"]),
        ("服务目录/SLA/监控评价", ["服务水平协议", "服务目录", "监控评价"]),
        ("重大信息科技项目", ["重大信息科技项目"]),
        ("数据处理活动及数据", ["数据处理", "数据安全", "个人信息", "重要数据"]),
        ("AI应用/模型/算法", ["人工智能", "模型", "算法"]),
        ("业务连续性/应急", ["业务连续性", "应急", "灾备"]),
        ("访问权限及信息资产", ["访问授权", "访问权限", "信息资产"]),
    ]
    found = [name for name, words in groups if any(word in text for word in words)]
    return "；".join(found[:4]) if found else "条款所述管理事项（需人工细化对象）"


def clause_timing(text: str) -> str:
    patterns = [
        r"至少每[^，。；]{1,12}", r"每年[^，。；]{0,12}", r"每三年[^，。；]{0,12}",
        r"合同签订前[^，。；]{0,15}", r"服务结束后[^，。；]{0,15}", r"到期前[^，。；]{0,15}",
        r"终止[^，。；]{0,12}前", r"定期", r"持续", r"及时", r"事先", r"24小时内",
    ]
    found = []
    for pattern in patterns:
        for value in re.findall(pattern, text):
            if value not in found:
                found.append(value)
    return "；".join(found[:5]) if found else "条款未明示固定时点/频率，按事项发生时及行内流程执行"


def clause_evidence(text: str) -> str:
    evidence = []
    pairs = [
        ("报告", "报告及报送/审议记录"), ("审批", "审批记录"), ("审核", "审核意见或会议结论"),
        ("评估", "评估表/评估报告"), ("合同", "合同、附件及审核版本"),
        ("协议", "协议及版本记录"), ("监控", "监控数据和评价记录"),
        ("演练", "演练计划、记录、报告及整改"), ("培训", "培训材料、签到及考核记录"),
        ("日志", "日志及留存校验记录"), ("清单", "清单和更新记录"),
    ]
    for keyword, item in pairs:
        if keyword in text and item not in evidence:
            evidence.append(item)
    if evidence:
        return "；".join(evidence[:5])
    return "条款未明示固定载体；为证明履职，应留存执行、复核、问题整改和关闭证据"


def make_clause(source_id: str, clause_id: str, text: str, title: str = "") -> Clause:
    value = compact(text)
    action = value[:220] + ("……" if len(value) > 220 else "")
    return Clause(source_id, clause_id, title, value, clause_subject(value), clause_object(value), action,
                  clause_timing(value), clause_evidence(value))


def split_articles(text: str, source_id: str) -> list[Clause]:
    normalized = compact_pdf_text(text) if "\n第\n" in text or text.count("\n") > len(text) / 4 else compact(text)
    pattern = re.compile(r"第([零〇一二两三四五六七八九十百千\d]+)条")
    matches = list(pattern.finditer(normalized))
    result = []
    seen = set()
    for index, match in enumerate(matches):
        number = cn_to_int(match.group(1))
        if number in seen:
            continue
        seen.add(number)
        end = matches[index + 1].start() if index + 1 < len(matches) else len(normalized)
        body = normalized[match.end():end].strip(" ：:\n")
        result.append(make_clause(source_id, f"第{number}条", body))
    return result


def plan_quoted_headings() -> list[str]:
    workbook = load_workbook(PLAN, data_only=True, read_only=True)
    sheet = workbook.active
    headings = []
    for row in range(2, sheet.max_row + 1):
        basis = str(sheet.cell(row, 12).value or "")
        for heading in re.findall(r"“([^”]+)”", basis):
            for item in re.split(r"[、]", heading):
                item = item.strip().replace("／", "/")
                if 2 <= len(item) <= 40 and item not in headings:
                    headings.append(item)
    return headings


def split_sections(text: str, source_id: str) -> list[Clause]:
    """Structure unnumbered internal rules by stable section headings."""
    positions = []
    for heading in list(dict.fromkeys(INTERNAL_SECTION_HEADINGS + plan_quoted_headings())):
        for match in re.finditer(re.escape(heading), text):
            positions.append((match.start(), match.end(), heading))
    # Prefer the longest heading when one heading is nested inside another, e.g. “评价管理”
    # inside “外包商评价管理”.
    positions.sort(key=lambda item: (item[0], -(item[1] - item[0])))
    non_overlapping = []
    last_end = -1
    for position in positions:
        if position[0] < last_end:
            continue
        non_overlapping.append(position)
        last_end = position[1]
    # Keep the first non-overlapping occurrence of a heading in document order.
    deduplicated = []
    seen = set()
    for position in non_overlapping:
        if position[2] not in seen:
            seen.add(position[2])
            deduplicated.append(position)
    deduplicated.sort()
    result = []
    for index, (start, body_start, heading) in enumerate(deduplicated):
        end = deduplicated[index + 1][0] if index + 1 < len(deduplicated) else len(text)
        body = compact(text[body_start:end]).strip(" ：:\n")
        if body:
            result.append(make_clause(source_id, f"章节:{heading}", body, heading))
    return result


def extract_markdown_articles(path: Path, source_id: str) -> list[Clause]:
    text = path.read_text(encoding="utf-8")
    pattern = re.compile(r"\*\*第([零〇一二两三四五六七八九十百千\d]+)条\*\*")
    matches = list(pattern.finditer(text))
    result = []
    for index, match in enumerate(matches):
        number = cn_to_int(match.group(1))
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        result.append(make_clause(source_id, f"第{number}条", text[match.end():end].strip()))
    return result


def extract_ai_items(path: Path) -> list[Clause]:
    text = path.read_text(encoding="utf-8")
    pattern = re.compile(r"^###\s+（([零〇一二两三四五六七八九十百千]+)）([^\n]+)", re.M)
    matches = list(pattern.finditer(text))
    result = []
    for index, match in enumerate(matches):
        number = cn_to_int(match.group(1))
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        result.append(make_clause("AI", f"第{number}项", text[match.end():end].strip(), match.group(2).strip()))
    return result


def extracted_xlsx_documents(path: Path) -> list[tuple[str, str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    result = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name = str(row[1] or row[0] or "未命名文件")
            status = str(row[4] or "")
            if "成功" not in status:
                continue
            text = "\n".join(str(value) for value in row[5:] if value)
            if text:
                result.append((name, text))
    return result


def build_clauses() -> list[Clause]:
    clauses: list[Clause] = list(MANUAL_CLAUSES)
    clauses.extend(extract_markdown_articles(REG_DIR / "银行保险机构信息科技外包风险监管办法_银保监办发2021_141号.md", "141"))
    clauses.extend(extract_ai_items(REG_DIR / "银行业保险业人工智能安全开发应用指导意见_金发2026_8号.md"))
    data_text = extract_pdf_text(REG_DIR / "国家金融监督管理总局关于印发银行保险机构数据安全管理办法的通知.pdf")
    clauses.extend(split_articles(data_text, "DATA"))

    # Parse the current internal master regulation; attachments are kept as separate source documents.
    for name, text in extracted_xlsx_documents(INTERNAL_3):
        if "风险管理办法（3.0版" in name and name.endswith((".docx", ".doc")):
            clauses.extend(split_articles(text, "GZ3"))
            break

    # Structure each successfully extracted supporting rule under a stable derived source id.
    for name, text in extracted_xlsx_documents(INTERNAL_RULES):
        if not any(key in name for key in ["供应商管理细则", "驻场服务管理细则", "通用外包管理细则", "外包管理规范", "资源使用管理细则", "专项应急预案", "框架协议外包管理"]):
            continue
        short = re.sub(r"\.[^.]+$", "", name)
        source_id = "GZ_RULES:" + short[:36]
        extracted = split_articles(text, source_id)
        if not extracted:
            extracted = split_sections(text, source_id)
        if not extracted:
            extracted = [make_clause(source_id, "全文/章节", text[:8000], short)]
        clauses.extend(extracted)

    # Apply authoritative semantics to selected 141 clauses.
    for clause in clauses:
        if clause.source_id == "141":
            match = re.search(r"(\d+)", clause.clause_id)
            if match and int(match.group(1)) in MANUAL_141:
                clause.subject, clause.object_, clause.action, clause.timing, clause.evidence = MANUAL_141[int(match.group(1))]
                clause.extraction_status = "原文自动结构化+人工语义校核"
    return clauses


def reference_numbers(segment: str, unit: str) -> list[int]:
    result: list[int] = []
    range_pattern = re.compile(rf"第?([零〇一二两三四五六七八九十百千\d]+){unit}?\s*(?:至|—|-)\s*第?([零〇一二两三四五六七八九十百千\d]+){unit}")
    for start, end in range_pattern.findall(segment):
        a, b = cn_to_int(start), cn_to_int(end)
        result.extend(range(a, b + 1))
    cleaned = range_pattern.sub("", segment)
    grouped = re.compile(rf"第([零〇一二两三四五六七八九十百千\d]+(?:[、，,][零〇一二两三四五六七八九十百千\d]+)+){unit}")
    for values in grouped.findall(cleaned):
        result.extend(cn_to_int(value) for value in re.split(r"[、，,]", values))
    cleaned = grouped.sub("", cleaned)
    for value in re.findall(rf"第([零〇一二两三四五六七八九十百千\d]+){unit}", cleaned):
        result.append(cn_to_int(value))
    return list(dict.fromkeys(result))


def detect_source(segment: str) -> str | None:
    for source_id, aliases in SOURCE_ALIASES.items():
        if any(alias in segment for alias in aliases):
            return source_id
    if "驻场服务管理细则" in segment:
        return "GZ_RULES:驻场"
    if "供应商管理细则" in segment:
        return "GZ_RULES:供应商"
    if "外包服务专项应急预案" in segment:
        return "GZ_RULES:专项应急预案"
    if "外包管理规范" in segment:
        return "GZ_RULES:外包管理规范"
    if "通用外包资源使用管理细则" in segment:
        return "GZ_RULES:通用外包资源使用"
    if "通用外包管理细则" in segment:
        return "GZ_RULES:通用外包管理"
    if "框架协议外包管理实施细则" in segment:
        return "GZ_RULES:框架协议外包管理"
    return None


def find_clause(index: dict[tuple[str, str], Clause], source_id: str, number: int, unit: str = "条") -> Clause | None:
    clause = index.get((source_id, f"第{number}{unit}"))
    if clause:
        return clause
    if source_id.startswith("GZ_RULES:"):
        keyword = source_id.split(":", 1)[1]
        for (candidate_source, clause_id), value in index.items():
            if candidate_source.startswith("GZ_RULES:") and keyword in candidate_source and clause_id == f"第{number}{unit}":
                return value
    return None


def find_section_clause(index: dict[tuple[str, str], Clause], source_id: str, heading: str) -> Clause | None:
    def normalized(value: str) -> str:
        return re.sub(r"\s+", "", value).replace("／", "/").replace("、", "/")

    keyword = source_id.split(":", 1)[1] if ":" in source_id else ""
    for (candidate_source, clause_id), value in index.items():
        if not candidate_source.startswith("GZ_RULES:"):
            continue
        if keyword and keyword not in candidate_source:
            continue
        if normalized(clause_id) == normalized(f"章节:{heading}"):
            return value
    return None


def resolve_basis(basis: str, index: dict[tuple[str, str], Clause]) -> tuple[list[Clause], list[str]]:
    resolved: list[Clause] = []
    unresolved: list[str] = []
    for segment in re.split(r"[；;]", basis or ""):
        segment = segment.strip()
        if not segment:
            continue
        source_id = detect_source(segment)
        if source_id == "PROC":
            continue
        if source_id is None:
            if any(word in segment for word in ["依据", "制度", "细则", "监管", "管理要求", "章节"]):
                unresolved.append(segment)
            continue
        unit = "项" if source_id == "AI" else "条"
        numbers = reference_numbers(segment, unit)
        if source_id.startswith("GZ_RULES:") and not numbers:
            headings = re.findall(r"“([^”]+)”", segment)
            # Normalize descriptive citations to the stable section name.
            if not headings:
                for heading in INTERNAL_SECTION_HEADINGS:
                    if heading in segment:
                        headings.append(heading)
            for heading_group in headings:
                for heading in re.split(r"[、，,]", heading_group):
                    normalized = heading.strip().replace("／", "/")
                    if normalized == "人员入场管理要求":
                        normalized = "入场管理"
                    clause = find_section_clause(index, source_id, normalized)
                    if clause:
                        resolved.append(clause)
                    else:
                        unresolved.append(f"{source_id}章节:{normalized}")
            if headings:
                continue
        if not numbers:
            unresolved.append(segment)
            continue
        for number in numbers:
            clause = find_clause(index, source_id, number, unit)
            if clause:
                resolved.append(clause)
            else:
                unresolved.append(f"{source_id}第{number}{unit}")
    return resolved, list(dict.fromkeys(unresolved))


def shorten(value: str, limit: int = 180) -> str:
    value = compact(value).replace("\n", " ")
    return value if len(value) <= limit else value[:limit] + "……"


def aggregate(clauses: Iterable[Clause], field: str, limit_items: int = 5) -> str:
    values = []
    for clause in clauses:
        value = getattr(clause, field)
        item = f"{clause.source_id}{clause.clause_id}：{shorten(value)}"
        if item not in values:
            values.append(item)
    return "\n".join(values[:limit_items]) if values else "当前依据未解析出可核验条款，需回到制度源文件确认。"


def basic_review(task_id: int, task: str, basis: str, applicability: str, evidence: str,
                 clauses: list[Clause], unresolved: list[str]) -> dict[str, str]:
    source_ids = {clause.source_id for clause in clauses}
    has_project = any(word in basis for word in ["采购文件", "采购交付", "付款条件", "合同"])
    generic = any(word in basis for word in ["外部监管适用条款", "AI指导意见。", "相关制度", "管理要求", "某章节"])
    ranges = bool(re.search(r"第?[零一二三四五六七八九十百千\d]+条?\s*(?:至|—|-)\s*第?[零一二三四五六七八九十百千\d]+条", basis))

    if task_id in REVIEW_PASS_OVERRIDES:
        return {"relation": REVIEW_PASS_OVERRIDES[task_id], "conclusion": "初审未发现明显错引",
                "issue": "无明显问题/方法性引用", "advice": "暂保留，后续按制度版本变化更新。",
                "direct": basis, "support": "无单独降级建议", "manual": "最终确认制度效力状态。",
                "rule": "R10-方法性任务或精确组合引用"}

    if task_id in REVIEW_OVERRIDES:
        issue, advice, direct, support, manual = REVIEW_OVERRIDES[task_id]
        relation = "存在问题：直接依据、支持依据或触发条件需要重新分层。"
        conclusion = "需调整制度依据表达（本轮只形成问题清单，不修改执行计划）"
        return {"relation": relation, "conclusion": conclusion, "issue": issue, "advice": advice,
                "direct": direct, "support": support, "manual": manual, "rule": "R8-已确认重点问题覆盖"}

    if unresolved:
        issue = "制度来源或条款未能精确解析"
        advice = "回到源文件核对条款原文；未核验前不得认定为直接依据。"
        direct = "待源条款核验后确定"
        support = "现有引用仅可暂列待核"
        manual = "；".join(unresolved[:6])
        relation = "无法确认直接性：当前引用存在章节化、概括化或源条款未结构化问题。"
        conclusion = "需人工核验"
        return {"relation": relation, "conclusion": conclusion, "issue": issue, "advice": advice,
                "direct": direct, "support": support, "manual": manual, "rule": "R2-不可解析引用进入人工校核门"}

    if generic or ranges:
        issue = "引用范围过宽或未精确到条款"
        advice = "将每个管理动作映射到具体条/款/项，并区分直接依据和支持依据。"
        direct = "待逐条拆分后确定"
        support = basis
        manual = "核对宽泛引用中是否包含本项目不适用条款。"
        relation = "只能说明总体相关，不能证明条款可直接推出本任务。"
        conclusion = "需精确化"
        return {"relation": relation, "conclusion": conclusion, "issue": issue, "advice": advice,
                "direct": direct, "support": support, "manual": manual, "rule": "R3-宽泛引用不得作为直接依据"}

    if has_project and not clauses:
        return {"relation": "属于采购/合同直接要求，不等同于监管条款直接义务。", "conclusion": "项目依据可保留",
                "issue": "无明显制度错引；需避免标为监管直接义务", "advice": "保留为采购/合同直接依据。",
                "direct": basis, "support": "无", "manual": "确认正式合同与采购文件版本。", "rule": "R6-项目依据独立分层"}

    if clauses:
        relation = "已解析到具体条款；从主体、对象、动作、时点和证据看，可作为直接或承接依据，仍需业务复核。"
        conclusion = "初审未发现明显错引"
        if "2009" in source_ids and any(clause.extraction_status.startswith("自动") for clause in clauses if clause.source_id == "2009"):
            conclusion = "需核2009年原文"
        return {"relation": relation, "conclusion": conclusion, "issue": "无明显问题/待业务确认",
                "advice": "暂保留；最终确认时核对条款原文与实际执行材料。", "direct": basis,
                "support": "无单独降级建议", "manual": "核对本项目实际触发条件和责任主体。", "rule": "R7-五要素一致性初审"}

    return {"relation": "当前依据不足以判断与任务的直接关系。", "conclusion": "需补充依据",
            "issue": "依据不足", "advice": "补充具体制度条款或明确其属于内部项目管理动作。",
            "direct": "待补充", "support": basis, "manual": "确认任务来源。", "rule": "R1-无可核验依据"}


def review_plan(clauses: list[Clause]) -> list[dict]:
    index = {(clause.source_id, clause.clause_id): clause for clause in clauses}
    workbook = load_workbook(PLAN, data_only=True)
    sheet = workbook.active
    result = []
    for row in range(2, sheet.max_row + 1):
        task_id = int(sheet.cell(row, 1).value)
        stage = str(sheet.cell(row, 2).value or "")
        task = str(sheet.cell(row, 4).value or "")
        attribute = str(sheet.cell(row, 5).value or "")
        applicability = str(sheet.cell(row, 6).value or "")
        owner = str(sheet.cell(row, 7).value or "")
        evidence = str(sheet.cell(row, 8).value or "")
        deadline = str(sheet.cell(row, 9).value or "")
        basis = str(sheet.cell(row, 12).value or "")
        resolved, unresolved = resolve_basis(basis, index)
        opinion = basic_review(task_id, task, basis, applicability, evidence, resolved, unresolved)
        result.append({
            "task_id": task_id, "stage": stage, "task": task, "attribute": attribute,
            "current_owner": owner, "current_evidence": evidence, "current_deadline": deadline,
            "current_basis": basis,
            "q1_subject": aggregate(resolved, "subject"),
            "q2_object": aggregate(resolved, "object_"),
            "q3_action": aggregate(resolved, "action"),
            "q4_timing": aggregate(resolved, "timing"),
            "q5_evidence": aggregate(resolved, "evidence") + "\n任务当前拟留存：" + evidence,
            "q6_applicability": applicability or "未标明，需补充直接/条件/机构层面适用判断。",
            "q7_relation": opinion["relation"],
            "conclusion": opinion["conclusion"], "issue_type": opinion["issue"],
            "advice": opinion["advice"], "suggested_direct": opinion["direct"],
            "suggested_support": opinion["support"], "manual_check": opinion["manual"],
            "resolved_clauses": "；".join(f"{c.source_id}{c.clause_id}" for c in resolved),
            "unresolved": "；".join(unresolved), "rule_hit": opinion["rule"],
        })
    return result


def style_sheet(sheet, freeze: str = "A2", filter_: bool = True) -> None:
    sheet.freeze_panes = freeze
    if filter_:
        sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 34
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def write_workbook(sources: list[Source], clauses: list[Clause], reviews: list[dict], output: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "00_复核说明与统计"
    issues = [row for row in reviews if row["conclusion"] not in ["初审未发现明显错引", "项目依据可保留"]]
    counts = {}
    for row in reviews:
        counts[row["conclusion"]] = counts.get(row["conclusion"], 0) + 1
    summary_rows = [
        ["项目", "内容"],
        ["工作流版本", "v1.3（已用当前223项数据完成多轮回放、无条号章节结构化及重点问题覆盖）"],
        ["复核范围", f"执行计划{len(reviews)}项任务；制度源{len(sources)}类；结构化条款{len(clauses)}条/项"],
        ["本轮边界", "只形成制度依据专项复核底稿和问题清单，不修改原执行计划。"],
        ["通过口径", "必须同时核对主体、对象、动作、时点/频率、证据、适用性以及直接/支持关系。"],
        ["人工校核门", "2009年指引除已确认第32、45、62条外，引用其他条款必须回看原文；征求意见稿不得作为现行直接依据。"],
        ["问题任务数", len(issues)],
    ]
    for key, value in sorted(counts.items()):
        summary_rows.append(["复核结论统计", f"{key}：{value}项"])
    for row in summary_rows:
        summary.append(row)
    style_sheet(summary, filter_=False)
    set_widths(summary, [20, 110])

    source_sheet = workbook.create_sheet("01_制度源文件清单")
    source_headers = ["来源ID", "制度名称", "文号", "层级", "效力状态", "生效/发布日期", "源文件", "格式", "结构化状态", "复核用途", "说明"]
    source_sheet.append(source_headers)
    for source in sources:
        source_sheet.append([source.source_id, source.title, source.document_no, source.level, source.effect_status,
                             source.effective_date, source.path, source.format, source.extraction_status,
                             source.review_use, source.note])
    style_sheet(source_sheet)
    set_widths(source_sheet, [13, 34, 22, 16, 18, 15, 62, 15, 32, 36, 55])

    clause_sheet = workbook.create_sheet("02_制度条款结构化")
    clause_headers = ["来源ID", "条款/项目", "标题", "条款原文", "责任主体", "管理对象", "强制动作", "时点/频率", "证据要求", "结构化状态"]
    clause_sheet.append(clause_headers)
    for clause in clauses:
        clause_sheet.append([clause.source_id, clause.clause_id, clause.title, clause.text, clause.subject,
                             clause.object_, clause.action, clause.timing, clause.evidence, clause.extraction_status])
    style_sheet(clause_sheet)
    set_widths(clause_sheet, [25, 12, 28, 90, 32, 34, 70, 30, 45, 28])

    review_sheet = workbook.create_sheet("03_223项逐项复核")
    review_headers = ["序号", "阶段", "执行事项", "任务属性", "当前制度依据", "1.条款要求谁做", "2.对什么对象做",
                      "3.必须做什么动作", "4.时点或频率", "5.必须留下什么证据", "6.本项目适用性",
                      "7.条款与任务关系", "复核结论", "问题类型", "建议处置", "建议直接依据",
                      "建议支持依据", "待人工核验", "已解析条款", "未解析引用", "命中规则"]
    review_sheet.append(review_headers)
    keys = ["task_id", "stage", "task", "attribute", "current_basis", "q1_subject", "q2_object", "q3_action",
            "q4_timing", "q5_evidence", "q6_applicability", "q7_relation", "conclusion", "issue_type",
            "advice", "suggested_direct", "suggested_support", "manual_check", "resolved_clauses", "unresolved", "rule_hit"]
    for review in reviews:
        review_sheet.append([review[key] for key in keys])
    style_sheet(review_sheet)
    set_widths(review_sheet, [8, 20, 42, 17, 60, 44, 42, 72, 40, 65, 48, 52, 24, 30, 62, 46, 46, 55, 34, 40, 30])

    issue_sheet = workbook.create_sheet("04_问题清单")
    issue_headers = ["序号", "阶段", "执行事项", "当前制度依据", "复核结论", "问题类型", "建议处置",
                     "建议直接依据", "建议支持依据", "适用性", "待人工核验", "命中规则"]
    issue_sheet.append(issue_headers)
    for review in issues:
        issue_sheet.append([review["task_id"], review["stage"], review["task"], review["current_basis"],
                            review["conclusion"], review["issue_type"], review["advice"],
                            review["suggested_direct"], review["suggested_support"], review["q6_applicability"],
                            review["manual_check"], review["rule_hit"]])
    style_sheet(issue_sheet)
    set_widths(issue_sheet, [8, 20, 42, 65, 24, 30, 65, 48, 48, 50, 55, 30])

    rules_sheet = workbook.create_sheet("05_判定规则与反馈")
    rules_sheet.append(["规则编号", "判定规则", "处置", "本轮数据反馈及修正"])
    rules = [
        ("R1", "无具体条款且不是明确采购/合同来源", "补充依据或说明为内部项目管理动作", "防止为了填满依据而牵强引用监管条款。"),
        ("R2", "源文件未结构化、条款无法解析或2009年未核原文", "进入人工校核门", "首次回放发现历史指引错引风险高，因此不得依靠编号推断。"),
        ("R3", "使用整份文件、章节或连续多条宽泛引用", "拆成原子义务", "第124项证明第23至29条整体引用会混入关联外包等不适用要求。"),
        ("R4", "责任主体、对象、动作、时点、证据任一不一致", "不得列作直接依据", "由SLA第四十五条错引抽象形成的五要素校验。"),
        ("R5", "条款仅支持任务中的部分动作", "降为支持依据或拆任务", "阶段评审、付款门与第24/25条的关系属于典型部分支持。"),
        ("R6", "任务由采购文件、合同、项目计划直接产生", "独立列项目依据", "不将采购付款、D01—D42基线伪装成监管直接义务。"),
        ("R7", "适用性属于条件或机构层面", "显式记录触发条件和项目配合边界", "业务连续性、重大项目、数据及AI条款必须设置触发判断。"),
        ("R8", "已确认的高风险任务", "用人工复核意见覆盖自动判断", "当前回放已覆盖SLA、合同、数据、AI、阶段评审、付款门、业务连续性、集中度、档案及2009年条款待核等重点任务。"),
        ("R9", "复核意见准备写回执行计划", "须经条款原文复核和业务事实确认", "本工作流先出问题清单，未经确认不自动改L列。"),
    ]
    for row in rules:
        rules_sheet.append(row)
    style_sheet(rules_sheet)
    set_widths(rules_sheet, [12, 48, 45, 80])

    feedback_sheet = workbook.create_sheet("06_人工确认回写区")
    feedback_sheet.append(["序号", "人工确认结论", "最终直接依据", "最终支持依据", "删除引用", "补充引用", "确认人/日期", "备注"])
    for review in reviews:
        feedback_sheet.append([review["task_id"], "待确认", "", "", "", "", "", ""])
    style_sheet(feedback_sheet)
    set_widths(feedback_sheet, [8, 22, 55, 55, 45, 45, 22, 50])

    workbook.save(output)


def write_json(sources: list[Source], clauses: list[Clause], reviews: list[dict]) -> None:
    DATA_DIR.mkdir(exist_ok=True)
    (DATA_DIR / "制度源文件清单.json").write_text(json.dumps([asdict(s) for s in sources], ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "制度条款结构化.json").write_text(json.dumps([asdict(c) for c in clauses], ensure_ascii=False, indent=2), encoding="utf-8")
    (DATA_DIR / "执行计划逐项复核.json").write_text(json.dumps(reviews, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", type=Path, default=OUTPUT)
    args = parser.parse_args()
    clauses = build_clauses()
    reviews = review_plan(clauses)
    write_workbook(SOURCE_SPECS, clauses, reviews, args.output)
    write_json(SOURCE_SPECS, clauses, reviews)
    issues = sum(row["conclusion"] not in ["初审未发现明显错引", "项目依据可保留"] for row in reviews)
    print(json.dumps({"output": str(args.output), "sources": len(SOURCE_SPECS), "clauses": len(clauses),
                      "tasks": len(reviews), "issues": issues}, ensure_ascii=False))


if __name__ == "__main__":
    main()
